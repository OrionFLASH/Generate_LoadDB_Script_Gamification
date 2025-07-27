#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Главный модуль для генерации JavaScript скриптов
Автор: OrionFLASH
Описание: Программа для создания JavaScript скриптов на основе входных данных
         с поддержкой логирования, измерения времени выполнения и копирования в буфер обмена
"""

import logging
import os
import time
import datetime
import csv
import re
import json
import pandas as pd
from functools import wraps

# Импорт библиотеки для работы с буфером обмена (удалено - не используется)
# import pyperclip

# Импорт библиотек для работы с Excel
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule

# =============================================================================
# ГЛОБАЛЬНЫЕ НАСТРОЙКИ ПРОГРАММЫ
# =============================================================================

# Базовая папка проекта
BASE_DIR = r"/Users/orionflash/Desktop/MyProject/Gen_Load_Game_Script/WORK"

# Настройки логирования
LOG_LEVEL = "INFO"  # Уровень детализации логов: "INFO" - основная информация, "DEBUG" - подробная отладочная информация
LOG_FILENAME_BASE = "LOG_2"  # Базовое имя файла лога (к нему добавляется дата и время)

# Имена подпапок (глобально)
SUBDIRECTORIES = {
    "LOGS": "LOGS",           # Папка для логов
    "INPUT": "INPUT",         # Папка для входных файлов
    "OUTPUT": "OUTPUT",       # Папка для выходных файлов
    "SCRIPT": "SCRIPT",       # Папка для сгенерированных скриптов
    "CONFIG": "CONFIG",       # Папка для конфигурационных файлов
    "JSON": "JSON"            # Папка для JSON файлов
}

# Расширения файлов для различных форматов (глобально)
FILE_EXTENSIONS = {
    "CSV": ".csv",    # Ключ: формат CSV файлов
    "TXT": ".txt",    # Ключ: формат текстовых файлов  
    "JSON": ".json",  # Ключ: формат JSON файлов
    "EXCEL": ".xlsx"  # Ключ: формат Excel файлов
}

# Выбор активных скриптов для генерации (глобально)
ACTIVE_SCRIPTS = [
    "leaders_for_admin",  # Скрипт для получения информации по участникам турнира (LeadersForAdmin)
    "reward",  # Скрипт для выгрузки профилей участников по кодам наград (Reward)
    # "reward_profiles",  # Обработка профилей наград из JSON в Excel (теперь внутри reward)
    # "reward",             # Скрипт для получения информации о наградах сотрудников
    # "profile",            # Скрипт для получения профилей сотрудников
    # "news_details",       # Скрипт для получения детальной карточки новости
    # "address_book_tn",    # Скрипт для получения карточки сотрудника по табельному номеру
    # "address_book_dev",   # Скрипт для получения карточки подразделения
    # "orders",             # Скрипт для получения списка сотрудников с преференциями
    # "news_list",          # Скрипт для получения списка новостей
    # "rating_list"         # Скрипт для получения рейтинга участников
]



# Названия листов для экспорта Excel
# Словарь с названиями листов, которые будут созданы в Excel файле
EXCEL_SHEET_NAMES = {
    "data": "DATA",       # Ключ: основной лист с данными (плоская структура из JSON)
    "summary": "SUMMARY", # Ключ: лист с общей сводкой и метаданными
    "statistics": "STATISTICS"  # Ключ: лист со статистическими данными и аналитикой
}

# Цвета для оформления Excel
# Словарь с HEX-кодами цветов для оформления Excel файлов
# Используется для создания профессионального внешнего вида отчетов
EXCEL_COLORS = {
    "header": "366092",     # Ключ: темно-синий цвет для заголовков (основной)
    "subheader": "9BC2E6",  # Ключ: светло-синий цвет для подзаголовков
    "alternate": "E7E6E6",  # Ключ: светло-серый цвет для чередующихся строк
    "highlight": "FFEB9C"   # Ключ: желтый цвет для выделения важных данных
}

# Настройки для TXT файлов
# Массив всех возможных разделителей для текстовых файлов
# Программа автоматически определяет разделитель, анализируя содержимое файла
# Включает знаки препинания, пробелы, переносы строк и специальные символы
TXT_DELIMITERS = [",", ";", "\t", " ", "\n", "\r\n", "|", ":", ".", "!", "?", "@", "#", "$", "%", "^", "&", "*", "(", ")", "[", "]", "{", "}", "<", ">", "/", "\\", "=", "+", "~", "`", "'", '"']

# Настройки для CSV файлов
CSV_DELIMITER = ";"  # Разделитель колонок в CSV файлах (точка с запятой для европейского формата)
CSV_ENCODING = "utf-8"  # Кодировка для CSV файлов (поддерживает кириллицу и специальные символы)
CSV_COLUMN_NAME = "data_column"  # Название столбца по умолчанию для извлечения данных (если не указан конкретный)

# Тестовые данные для работы без внешнего файла
# Список тестовых значений, используемых когда DATA_SOURCE = "variable"
# Позволяет тестировать программу без необходимости создания входных файлов
TEST_DATA_LIST = [
    "test_value_1",  # Тестовое значение 1
    "test_value_2",  # Тестовое значение 2
    "test_value_3"   # Тестовое значение 3
]

# =============================================================================
# ТЕКСТЫ ДЛЯ ЛОГИРОВАНИЯ
# =============================================================================

# Словарь с сообщениями для логирования
# Содержит все текстовые сообщения, выводимые программой
# Поддерживает форматирование с переменными в фигурных скобках {variable}
# Позволяет централизованно управлять всеми сообщениями программы
LOG_MESSAGES = {
    # Сообщения о начале и конце программы
    "program_start": "=== СТАРТ ПРОГРАММЫ - Генератор JavaScript скриптов: {time} ===",  # Ключ: сообщение о старте программы
    "program_end": "=== ФИНАЛ ПРОГРАММЫ - {time} ===",  # Ключ: сообщение о завершении программы
    "processing_start_time": "Время начала обработки: {time}",  # Ключ: время начала обработки
    "logging_level": "Уровень логирования: {level}",  # Ключ: уровень логирования
    "total_execution_time": "Итоговое время работы: {time:.4f} секунд",  # Ключ: общее время выполнения
    
    # Сообщения о выполнении функций
    "function_start": "[START] {func} {params}",  # Ключ: начало выполнения функции
    "function_completed": "[END] {func} {params} (время: {time}s)",  # Ключ: завершение функции
    "function_error": "[ERROR] {func} {params} — {error}",  # Ключ: ошибка в функции
    
    # Сообщения о данных
    "data_received": "Получено данных для обработки: {count}",  # Ключ: количество полученных данных
    "program_success": "Программа выполнена успешно",  # Ключ: успешное завершение программы
    "critical_error": "Критическая ошибка в программе: {error}",  # Ключ: критическая ошибка
    
    # Сообщения для итоговой статистики
    "summary_title": "SUMMARY - Итоговая статистика",  # Ключ: заголовок итоговой статистики
    "total_time": "Общее время: {time:.4f} сек",  # Ключ: общее время выполнения
    "actions_processed": "Действий: {count}",  # Ключ: количество обработанных действий
    "functions_executed": "Функций: {count}",  # Ключ: количество выполненных функций
    "function_time": "Функция {func}: {time:.4f} сек",  # Ключ: время выполнения конкретной функции
    "program_completed": "Программа завершена: {time}",  # Ключ: время завершения программы
    
    # Сообщения о работе с файлами
    "file_loading": "Загрузка данных из файла: {file_path}, формат: {format}",  # Ключ: загрузка файла
    "file_not_found": "Файл не найден: {file_path}",  # Ключ: файл не найден
    "file_loaded": "Файл успешно загружен: {file_path}, элементов: {count}",  # Ключ: файл загружен
    "file_load_error": "Ошибка загрузки файла: {file_path}. {error}",  # Ключ: ошибка загрузки файла
    "using_test_data": "Использование тестовых данных: {count} элементов",  # Ключ: использование тестовых данных
    
    # Сообщения о буфере обмена
    # Сообщения для буфера обмена удалены - функциональность не используется
    # "clipboard_copied": "Текст скопирован в буфер обмена",  # Ключ: успешное копирование в буфер
    # "clipboard_error": "Ошибка при копировании в буфер: {error}",  # Ключ: ошибка копирования в буфер
    
    # Сообщения о генерации скриптов
    "script_generation": "Генерация скрипта: {script_name}",  # Ключ: начало генерации скрипта
    "script_generated": "Скрипт {script_name} сгенерирован успешно (данных: {count})",  # Ключ: успешная генерация скрипта
    "script_saved": "Скрипт сохранен в файл: {file_path}",  # Ключ: скрипт сохранен в файл
    "script_save_error": "Ошибка сохранения скрипта: {error}",  # Ключ: ошибка сохранения скрипта
    
    # Сообщения для итоговой статистики
    "summary_stats": "ИТОГОВАЯ СТАТИСТИКА РАБОТЫ ПРОГРАММЫ",  # Ключ: заголовок статистики
    "total_execution": "Общее время выполнения: {time:.4f} секунд",  # Ключ: общее время выполнения
    "processed_actions": "Обработано действий: {count}",  # Ключ: количество обработанных действий
    "executed_functions": "Выполнено функций: {count}",  # Ключ: количество выполненных функций
    "execution_times": "Время выполнения функций:",  # Ключ: заголовок времени выполнения функций
    "selected_script": "Выбранный скрипт для генерации: {script_name}",  # Ключ: выбранный скрипт
    "config_loaded": "Конфигурация загружена для: {script_name}",  # Ключ: загрузка конфигурации
    
    # Сообщения о обработке файлов
    "csv_processing": "Обработка CSV: разделитель '{delimiter}', кодировка '{encoding}', столбец '{column}'",  # Ключ: обработка CSV файла
    "txt_processing": "Обработка TXT: найдено разделителей {delimiters_count}",  # Ключ: обработка TXT файла
    "data_source_selected": "Источник данных: {source} ({format})",  # Ключ: выбранный источник данных
    
    # Сообщения о обработке JSON файлов
    "json_conversion_start": "Начинаем конвертацию JSON: {input} -> {output}",  # Ключ: начало конвертации JSON
    "json_file_not_found": "JSON файл не найден: {file_path}",  # Ключ: JSON файл не найден
    "json_directory_created": "Создана директория: {directory}",  # Ключ: создана директория для JSON
    "json_data_loading": "Загружаем JSON данные...",  # Ключ: загрузка JSON данных
    "json_data_processing": "Обрабатываем JSON данные...",  # Ключ: обработка JSON данных
    "json_leaders_found": "Найдены данные лидеров в ключе: {key}, количество: {count}",  # Ключ: найдены данные лидеров
    "json_reward_found": "Найдены данные участников в коде: {key}, количество: {count}",  # Ключ: найдены данные участников наград
    "json_direct_leaders": "Прямой список лидеров, количество: {count}",  # Ключ: прямой список лидеров
    "json_invalid_format": "Неверный формат JSON данных",  # Ключ: неверный формат JSON
    "json_no_leaders": "Не найдены данные лидеров в JSON файле",  # Ключ: нет данных лидеров
    "json_records_processed": "Обработано {count} записей",  # Ключ: количество обработанных записей
    "json_excel_creation": "Создаем Excel файл...",  # Ключ: создание Excel файла
    "json_excel_success": "Excel файл успешно создан: {file_path}",  # Ключ: Excel файл создан
    
    # Сообщения о настройках колонок
    "column_settings_applying": "Применяем настройки колонок к DataFrame",  # Ключ: применение настроек колонок
    "column_settings_applied": "После применения настроек: {count} колонок",  # Ключ: результат применения настроек
    "column_settings_applying_leaders": "Применяем настройки колонок к DataFrame для leaders",  # Ключ: применение настроек для лидеров
    "no_active_scripts": "Нет активных скриптов для обработки. Настройте ACTIVE_SCRIPTS.",  # Ключ: нет активных скриптов
    "json_conversion_error": "Ошибка при конвертации JSON: {error}",  # Ключ: ошибка конвертации JSON
    "json_file_processing": "Обработка JSON файла: {file_name}",  # Ключ: обработка конкретного JSON файла
    "json_files_processed": "Обработано JSON файлов: {count}",  # Ключ: количество обработанных JSON файлов
    "json_no_files_found": "JSON файлы для обработки не найдены",  # Ключ: JSON файлы не найдены
    
    # Сообщения для обработки данных
    "float_conversion_error": "Ошибка преобразования '{val}' в float: {ex} | Context: {context}",  # Ключ: ошибка преобразования в float
    "reward_summary_sheet_created": "Лист REWARD_SUMMARY создан успешно",  # Ключ: лист наград создан
    "variant_selected": "Выбранный вариант: {variant}",  # Ключ: выбранный вариант
    "script_generation_start": "=== ГЕНЕРАЦИЯ СКРИПТА: {script_name} ===",  # Ключ: начало генерации скрипта
    "data_loading": "Загрузка данных и конфигурации...",  # Ключ: загрузка данных
    "config_loaded_count": "Конфигурация загружена: {count} элементов",  # Ключ: конфигурация загружена с количеством
    "domain_info": "Домен: {domain}",  # Ключ: информация о домене
    "api_path_info": "API путь: {api_path}",  # Ключ: информация о пути API
    "request_params": "Параметры запросов: delay={delay}, max_retries={max_retries}, timeout={timeout}",  # Ключ: параметры запросов
    "base_url_info": "Базовый URL: {base_url}",  # Ключ: базовый URL
    "ids_generated": "Строка IDs сгенерирована: {count} элементов",  # Ключ: IDs сгенерированы
    "script_saving": "Сохранение скрипта в файл...",  # Ключ: сохранение скрипта
    "script_generated_success": "Скрипт {script_name} сгенерирован успешно (данных: {count})",  # Ключ: скрипт сгенерирован успешно
    "json_load_error": "Ошибка при загрузке JSON файла {file_path}: {error}",  # Ключ: ошибка загрузки JSON
    "excel_creation_error": "Ошибка при создании Excel файла: {error}",  # Ключ: ошибка создания Excel
    "tournaments_processed": "Обработано турниров: {tournaments}, общее количество лидеров: {leaders}",  # Ключ: турниры обработаны
    "no_data_warning": "Нет данных для обработки",  # Ключ: нет данных
    "json_leaders_conversion_error": "Ошибка при конвертации JSON лидеров в Excel: {error}",  # Ключ: ошибка конвертации лидеров
    "profile_extraction_error": "Ошибка при извлечении профилей из данных: {error}",  # Ключ: ошибка извлечения профилей
    "reward_profiles_found": "Найдено профилей для кода награды {code}: {count} (структура: {structure})",  # Ключ: найдены профили наград
    "reward_profiles_found_old": "Найдено профилей для кода награды {code}: {count} (старая структура)",  # Ключ: найдены профили наград (старая структура)
    "rewards_processed": "Обработано кодов наград: {rewards}, общее количество профилей: {profiles}",  # Ключ: награды обработаны
    "direct_profiles_list": "Прямой список профилей: {count}",  # Ключ: прямой список профилей
    "no_profiles_error": "Не найдено данных профилей",  # Ключ: нет профилей
    "json_reward_conversion_error": "Ошибка при конвертации JSON наград в Excel: {error}",  # Ключ: ошибка конвертации наград
    "excel_file_creation": "Создаем Excel файл: {filename}",  # Ключ: создание Excel файла
    "separator_line": "=" * 70,  # Ключ: разделительная линия
    "active_scripts_info": "Активные скрипты: {scripts}",  # Ключ: активные скрипты
    "stage1_title": "=== ЭТАП 1: ГЕНЕРАЦИЯ СКРИПТОВ ===",  # Ключ: заголовок этапа 1
    "script_processing": "--- ОБРАБОТКА СКРИПТА: {script_name} ---",  # Ключ: обработка скрипта
    "active_operations_info": "Активные операции для {script_name}: {operations}",  # Ключ: активные операции
    "script_generation_info": "Генерация скрипта: {script_name}",  # Ключ: генерация скрипта
    "script_generation_skipped": "Пропуск генерации скрипта для {script_name} (режим: {operations})",  # Ключ: пропуск генерации
    "unknown_script_error": "Неизвестный скрипт: {script_name}",  # Ключ: неизвестный скрипт
    "stage2_title": "=== ЭТАП 2: ОБРАБОТКА JSON ФАЙЛОВ В EXCEL ===",  # Ключ: заголовок этапа 2
    "json_file_processing_info": "Обработка JSON файла: {json_file}",  # Ключ: обработка JSON файла
    "no_json_file_warning": "Для скрипта {script_name} не указан json_file",  # Ключ: нет JSON файла
    "json_processing_skipped": "Пропуск обработки JSON для {script_name} (режим: {operations})",  # Ключ: пропуск обработки JSON
    "numeric_conversion_error": "Ошибка преобразования в числовой формат колонки {column}, значение '{value}': {error}",  # Ключ: ошибка числового преобразования
    "date_conversion_error": "Ошибка преобразования в дату колонки {column}, значение '{value}': {error}",  # Ключ: ошибка преобразования даты
    "column_conversion_start": "Начинаем преобразование колонки {column} в тип {type}",  # Ключ: начало преобразования колонки
    "column_conversion_success": "Успешно преобразована колонка {column} в тип {type} ({action})",  # Ключ: успешное преобразование колонки
    "column_conversion_failed": "Ошибка преобразования колонки {column} в тип {type}: {error}",  # Ключ: ошибка преобразования колонки
    "group_conversion_start": "Начинаем преобразование группы {group} в тип {type} для полей: {fields}",  # Ключ: начало преобразования группы
    "group_conversion_completed": "Завершено преобразование группы {group}, обработано полей: {processed}",  # Ключ: завершение преобразования группы
    "column_not_found": "Колонка {column} не найдена в группе {group}",  # Ключ: колонка не найдена
    "columns_filtered": "{action} {count} колонок: {columns}",  # Ключ: фильтрация колонок
    "no_active_scripts": "Нет активных скриптов для обработки. Настройте ACTIVE_SCRIPTS.",  # Ключ: нет активных скриптов
    "summary_output": "Итоговая статистика работы программы:\n{summary}",  # Ключ: итоговая статистика
    "reward_profiles_leaders_found": "Найдены лидеры для кода награды {code}: {count} (структура: {structure})",  # Ключ: найдены лидеры наград
    "reward_profiles_leaders_processed": "Обработано кодов наград: {rewards}, общее количество лидеров: {leaders}",  # Ключ: лидеры наград обработаны
    "json_reward_profiles_conversion_error": "Ошибка при конвертации JSON профилей наград в Excel: {error}"  # Ключ: ошибка конвертации профилей наград
}

# =============================================================================
# КОНФИГУРАЦИЯ ФУНКЦИЙ
# =============================================================================

# Словарь с конфигурацией для каждого типа скрипта
# Содержит параметры для генерации JavaScript и обработки данных
# Каждая конфигурация включает: домен, API пути, параметры запросов, настройки файлов
FUNCTION_CONFIGS = {
    "leaders_for_admin": {  # Ключ: конфигурация для скрипта LeadersForAdmin (информация по участникам турнира)
        "name": "LeadersForAdmin",  # Ключ: название скрипта для отображения
        "description": "Информация по загруженным в турнир данным об участниках",  # Ключ: описание назначения скрипта
        "active_operations": "both",  # Ключ: активные операции ("scripts_only", "json_only", "both")
        "excel_freeze_row": 1,  # Ключ: номер строки для закрепления в Excel (1 = заголовок)
        "variants": {  # Ключ: варианты конфигурации (SIGMA/ALPHA)
            "sigma": {  # Ключ: вариант SIGMA (продакшн окружение)
                "name": "LeadersForAdmin (SIGMA)",  # Ключ: название варианта
                "description": "Информация по загруженным в турнир данным об участниках - SIGMA",  # Ключ: описание варианта
                "domain": "https://salesheroes.sberbank.ru",  # Ключ: домен для SIGMA
                "params": {  # Ключ: параметры API запросов
                    "api_path": "/bo/rmkib.gamification/api/v1/tournaments/",  # Ключ: путь к API
                    "service": "leadersForAdmin",  # Ключ: название сервиса
                    "page_param": "pageNum=1"  # Ключ: параметр пагинации
                }
            },
            "alpha": {  # Ключ: вариант ALPHA (тестовое окружение)
                "name": "LeadersForAdmin (ALPHA)",  # Ключ: название варианта
                "description": "Информация по загруженным в турнир данным об участниках - ALPHA",  # Ключ: описание варианта
                "domain": "https://efs-our-business-prom.omega.sbrf.ru",  # Ключ: домен для ALPHA
                "params": {  # Ключ: параметры API запросов
                    "api_path": "/bo/rmkib.gamification/api/v1/tournaments/",  # Ключ: путь к API
                    "service": "leadersForAdmin",  # Ключ: название сервиса
                    "page_param": "pageNum=1"  # Ключ: параметр пагинации
                }
            }
        },
        "timeout": 30000,  # Ключ: таймаут запроса в миллисекундах (общий для всех вариантов)
        "retry_count": 3,  # Ключ: количество попыток при ошибке (общий для всех вариантов)
        "delay_between_requests": 5,  # Ключ: задержка между запросами в секундах (общий для всех вариантов)
        "processing_options": {  # Ключ: опции обработки данных
            "remove_photo_data": True,  # Ключ: удалять ли поля photoData из результатов
            "include_division_ratings": True,  # Ключ: включать ли рейтинги подразделений
            "include_tournament_info": True  # Ключ: включать ли информацию о турнирах
        },
        "data_source": "external_file",  # Ключ: источник данных (file/variable/external_file)
        "input_format": "CSV",  # Ключ: формат входного файла
        "csv_column": "TOURNAMENT_CODE",  # Ключ: название столбца для извлечения данных
        "csv_delimiter": ";",  # Ключ: разделитель в CSV файле
        "csv_encoding": "utf-8",  # Ключ: кодировка CSV файла
        "input_file": "TOURNAMENT-SCHEDULE (PROM) 2025-07-25 v6",  # Ключ: имя входного файла (без расширения)
        "leaders_processing": {  # Ключ: конфигурация для обработки лидеров турниров (JSON → Excel)
            "name": "Leaders Processing",  # Ключ: название скрипта для отображения
            "description": "Обработка лидеров турниров из JSON в Excel",  # Ключ: описание назначения скрипта
            "active_operations": "json_only",  # Ключ: активные операции ("scripts_only", "json_only", "both")
            "excel_freeze_row": 1,  # Ключ: номер строки для закрепления в Excel (1 = заголовок)
            "json_file": "leadersForAdmin_SIGMA_20250727-130522",  # Ключ: имя JSON файла для обработки (без расширения)
            "excel_file": "LeadersForAdmin",  # Ключ: имя Excel файла для создания (без расширения)
            "excel_freeze_cell": "B2",  # Ключ: ячейка для закрепления в Excel (B2 = первая строка и первая колонка)
            "column_settings": {  # Ключ: настройки обработки колонок
                "columns_to_keep": [],  # Ключ: колонки для сохранения (если пусто - оставляем все)
                "columns_to_remove": [  # Ключ: колонки для удаления
                ],
                "numeric_conversions": {  # Ключ: преобразования в числовой формат
                    "float_fields": {  # Ключ: группа полей для преобразования в дробные числа
                        "fields": ["indicatorValue", "successValue"],  # Ключ: массив полей для преобразования
                        "type": "float",  # Ключ: тип числа ("integer" или "float")
                        "decimal_places": 2,  # Ключ: количество знаков после запятой
                        "replace_original": False  # Ключ: заменить исходное поле (True) или создать новое (False)
                    },
                    "integer_fields": {  # Ключ: группа полей для преобразования в целые числа
                        "fields": ["BANK_groupId", "TB_groupId", "GOSB_groupId", "BANK_placeInRating", "TB_placeInRating", "GOSB_placeInRating"],  # Ключ: массив полей для преобразования
                        "type": "integer",  # Ключ: тип числа ("integer" или "float")
                        "replace_original": True  # Ключ: заменить исходное поле (True) или создать новое (False)
                    }
                },
                "date_conversions": {  # Ключ: преобразования в формат даты (пока нет полей с датами в leaders)
                }
            }
        }
    },
    "reward": {  # Ключ: конфигурация для скрипта REWARD (выгрузка профилей участников по кодам наград)
        "name": "Reward",  # Ключ: название скрипта для отображения
        "description": "Выгрузка профилей участников по кодам наград",  # Ключ: описание назначения скрипта
        "active_operations": "both",  # Ключ: активные операции ("scripts_only", "json_only", "both")
        "variants": {  # Ключ: варианты конфигурации (SIGMA/ALPHA)
            "sigma": {  # Ключ: вариант SIGMA (продакшн окружение)
                "name": "Reward (SIGMA)",  # Ключ: название варианта
                "description": "Выгрузка профилей участников по кодам наград - SIGMA",  # Ключ: описание варианта
                "domain": "https://salesheroes.sberbank.ru",  # Ключ: домен для SIGMA
                "params": {
                    "api_path": "/bo/rmkib.gamification/api/v1/badges/",  # Ключ: путь к API
                    "service": "profiles"  # Ключ: название сервиса
                }
            },
            "alpha": {  # Ключ: вариант ALPHA (тестовое окружение)
                "name": "Reward (ALPHA)",  # Ключ: название варианта
                "description": "Выгрузка профилей участников по кодам наград - ALPHA",  # Ключ: описание варианта
                "domain": "https://efs-our-business-prom.omega.sbrf.ru",  # Ключ: домен для ALPHA
                "params": {
                    "api_path": "/bo/rmkib.gamification/api/v1/badges/",  # Ключ: путь к API
                    "service": "profiles"  # Ключ: название сервиса
                }
            }
        },
        "timeout": 30000,  # Ключ: таймаут запроса в миллисекундах (общий для всех вариантов)
        "retry_count": 3,  # Ключ: количество попыток при ошибке (общий для всех вариантов)
        "delay_between_requests": 5,  # Ключ: задержка между запросами в секундах (общий для всех вариантов)
        "data_source": "external_file",  # Ключ: источник данных (file/variable/external_file)
        "input_format": "CSV",  # Ключ: формат входного файла
        "csv_column": "REWARD_CODE",  # Ключ: название столбца для извлечения данных
        "csv_delimiter": ";",  # Ключ: разделитель в CSV файле
        "csv_encoding": "utf-8",  # Ключ: кодировка CSV файла
        "input_file": "REWARD (PROM) 2025-07-24 v1",  # Ключ: имя входного файла (без расширения)
        "processing_options": {  # Ключ: опции обработки данных
            "include_photo_data": False,  # Ключ: включать ли данные фотографий
            "remove_photo_data": True,  # Ключ: удалять ли поля photoData из результатов
            "include_division_ratings": True,  # Ключ: включать ли рейтинги подразделений
            "include_badge_info": True,  # Ключ: включать ли информацию о наградах
            "max_profiles_per_request": 1000,  # Ключ: максимальное количество профилей на запрос
            "skip_empty_profiles": True  # Ключ: пропускать ли пустые профили
        },
        "reward_profiles": {  # Ключ: конфигурация для обработки профилей наград (JSON → Excel)
            "name": "Reward Profiles",  # Ключ: название скрипта для отображения
            "description": "Обработка профилей наград из JSON в Excel",  # Ключ: описание назначения скрипта
            "active_operations": "json_only",  # Ключ: активные операции ("scripts_only", "json_only", "both")
            "excel_freeze_row": 1,  # Ключ: номер строки для закрепления в Excel (1 = заголовок)
            "json_file": "profiles_SIGMA_20250727-130833",  # Ключ: имя JSON файла для обработки (без расширения)
            "excel_file": "RewardProfiles",  # Ключ: имя Excel файла для создания (без расширения)
            "excel_freeze_cell": "F2",  # Ключ: ячейка для закрепления в Excel (B2 = первая строка и первая колонка)
            "column_settings": {  # Ключ: настройки обработки колонок
                "columns_to_keep": [],  # Ключ: колонки для сохранения (если пусто - оставляем все)
                "columns_to_remove": [  # Ключ: колонки для удаления
                    "isMarked", "colorPrimary", "colorSecondary",
                    "tag1_id", "tag1_name", "tag1_color",
                    "tag2_id", "tag2_name", "tag2_color", 
                    "tag3_id", "tag3_name", "tag3_color",
                    "tag4_id", "tag4_name", "tag4_color",
                    "tag5_id", "tag5_name", "tag5_color"
                ],
                "numeric_conversions": {  # Ключ: преобразования в числовой формат
                    "integer_fields": {  # Ключ: группа полей для преобразования в целые числа
                        "fields": ["gosbCode"],  # Ключ: массив полей для преобразования
                        "type": "integer",  # Ключ: тип числа ("integer" или "float")
                        "replace_original": True  # Ключ: заменить исходное поле (True) или создать новое (False)
                    }
                },
                "date_conversions": {  # Ключ: преобразования в формат даты
                    "receivingDate": {
                        "input_format": "DD.MM.YY",  # Ключ: входной формат даты
                        "output_format": "YYYY-MM-DD",  # Ключ: выходной формат даты
                        "replace_original": True  # Ключ: заменить исходное поле (True) или создать новое (False)
                    }
                }
            }
        }
    },
    "profile": {  # Ключ: конфигурация для скрипта PROFILE (профили сотрудников)
        "name": "PROFILE",  # Ключ: название скрипта для отображения
        "description": "Профили сотрудников в героях продаж",  # Ключ: описание назначения скрипта
        "active_operations": "scripts_only",  # Ключ: активные операции ("scripts_only", "json_only", "both")
        "domain": "profiles.example.com",  # Ключ: домен для API запросов
        "params": {  # Ключ: параметры API запросов
            "api_endpoint": "/api/profiles/employee",  # Ключ: конечная точка API
            "include_stats": True,  # Ключ: включать ли статистику
            "include_achievements": True,  # Ключ: включать ли достижения
            "format": "detailed"  # Ключ: формат ответа
        },
        "data_source": "file",  # Ключ: источник данных
        "input_format": "TXT",  # Ключ: формат входного файла
        "csv_column": "profile_id",  # Ключ: название столбца для извлечения данных
        "csv_delimiter": ";",  # Ключ: разделитель в CSV файле
        "csv_encoding": "utf-8",  # Ключ: кодировка CSV файла
        "json_file": "profile"  # Ключ: имя JSON файла для обработки (без расширения)
    },
    "news_details": {  # Ключ: конфигурация для скрипта NewsDetails (детальная карточка новости)
        "name": "NewsDetails",  # Ключ: название скрипта для отображения
        "description": "Детальная карточка новости",  # Ключ: описание назначения скрипта
        "active_operations": "scripts_only",  # Ключ: активные операции ("scripts_only", "json_only", "both")
        "domain": "news.example.com",  # Ключ: домен для API запросов
        "params": {  # Ключ: параметры API запросов
            "api_endpoint": "/api/news/details",  # Ключ: конечная точка API
            "include_content": True,  # Ключ: включать ли содержимое
            "include_attachments": True,  # Ключ: включать ли вложения
            "format": "full"  # Ключ: формат ответа
        },
        "data_source": "file",  # Ключ: источник данных
        "input_format": "TXT",  # Ключ: формат входного файла
        "csv_column": "news_id",  # Ключ: название столбца для извлечения данных
        "csv_delimiter": ";",  # Ключ: разделитель в CSV файле
        "csv_encoding": "utf-8",  # Ключ: кодировка CSV файла
        "json_file": "news_details"  # Ключ: имя JSON файла для обработки (без расширения)
    },
    "address_book_tn": {  # Ключ: конфигурация для скрипта AdressBookTN (карточка сотрудника по табельному номеру)
        "name": "AdressBookTN",  # Ключ: название скрипта для отображения
        "description": "Карточка сотрудника из адресной книги по табельным номерам",  # Ключ: описание назначения скрипта
        "active_operations": "scripts_only",  # Ключ: активные операции ("scripts_only", "json_only", "both")
        "domain": "directory.example.com",  # Ключ: домен для API запросов
        "params": {  # Ключ: параметры API запросов
            "api_endpoint": "/api/directory/employee",  # Ключ: конечная точка API
            "search_by": "employee_number",  # Ключ: поле для поиска
            "include_contacts": True,  # Ключ: включать ли контакты
            "include_department": True  # Ключ: включать ли отдел
        },
        "data_source": "file",  # Ключ: источник данных
        "input_format": "CSV",  # Ключ: формат входного файла
        "csv_column": "employee_number",  # Ключ: название столбца для извлечения данных
        "csv_delimiter": ";",  # Ключ: разделитель в CSV файле
        "csv_encoding": "utf-8",  # Ключ: кодировка CSV файла
        "json_file": "address_book_tn"  # Ключ: имя JSON файла для обработки (без расширения)
    },
    "address_book_dev": {  # Ключ: конфигурация для скрипта AdressBookDev (карточка подразделения)
        "name": "AdressBookDev",  # Ключ: название скрипта для отображения
        "description": "Карточка подразделения из адресной книги со списком сотрудников",  # Ключ: описание назначения скрипта
        "active_operations": "scripts_only",  # Ключ: активные операции ("scripts_only", "json_only", "both")
        "domain": "directory.example.com",  # Ключ: домен для API запросов
        "params": {  # Ключ: параметры API запросов
            "api_endpoint": "/api/directory/department",  # Ключ: конечная точка API
            "include_employees": True,  # Ключ: включать ли сотрудников
            "include_hierarchy": True,  # Ключ: включать ли иерархию
            "format": "detailed"  # Ключ: формат ответа
        },
        "data_source": "file",  # Ключ: источник данных
        "input_format": "CSV",  # Ключ: формат входного файла
        "csv_column": "department_id",  # Ключ: название столбца для извлечения данных
        "csv_delimiter": ";",  # Ключ: разделитель в CSV файле
        "csv_encoding": "utf-8",  # Ключ: кодировка CSV файла
        "json_file": "address_book_dev"  # Ключ: имя JSON файла для обработки (без расширения)
    },
    "orders": {  # Ключ: конфигурация для скрипта Orders (список сотрудников с преференциями)
        "name": "Orders",  # Ключ: название скрипта для отображения
        "description": "Список сотрудников выбравших преференции",  # Ключ: описание назначения скрипта
        "active_operations": "scripts_only",  # Ключ: активные операции ("scripts_only", "json_only", "both")
        "domain": "orders.example.com",  # Ключ: домен для API запросов
        "params": {  # Ключ: параметры API запросов
            "api_endpoint": "/api/orders/preferences",  # Ключ: конечная точка API
            "status": "selected",  # Ключ: статус заказов
            "include_details": True,  # Ключ: включать ли детали
            "date_from": "2024-01-01"  # Ключ: дата начала периода
        },
        "data_source": "file",  # Ключ: источник данных
        "input_format": "CSV",  # Ключ: формат входного файла
        "csv_column": "employee_id",  # Ключ: название столбца для извлечения данных
        "csv_delimiter": ";",  # Ключ: разделитель в CSV файле
        "csv_encoding": "utf-8",  # Ключ: кодировка CSV файла
        "json_file": "orders"  # Ключ: имя JSON файла для обработки (без расширения)
    },
    "news_list": {  # Ключ: конфигурация для скрипта NewsList (список новостей)
        "name": "NewsList",  # Ключ: название скрипта для отображения
        "description": "Список новостей",  # Ключ: описание назначения скрипта
        "active_operations": "scripts_only",  # Ключ: активные операции ("scripts_only", "json_only", "both")
        "domain": "news.example.com",  # Ключ: домен для API запросов
        "params": {  # Ключ: параметры API запросов
            "api_endpoint": "/api/news/list",  # Ключ: конечная точка API
            "status": "published",  # Ключ: статус новостей
            "include_preview": True,  # Ключ: включать ли превью
            "limit": 100  # Ключ: лимит записей
        },
        "data_source": "variable",  # Ключ: источник данных (использует тестовые данные)
        "input_format": "TXT",  # Ключ: формат входного файла
        "csv_column": "news_category",  # Ключ: название столбца для извлечения данных
        "csv_delimiter": ";",  # Ключ: разделитель в CSV файле
        "csv_encoding": "utf-8",  # Ключ: кодировка CSV файла
        "json_file": "news_list"  # Ключ: имя JSON файла для обработки (без расширения)
    },
    "rating_list": {  # Ключ: конфигурация для скрипта RaitingList (рейтинг участников)
        "name": "RaitingList",  # Ключ: название скрипта для отображения
        "description": "Рейтинг участников по полученным наградам и кристаллам",  # Ключ: описание назначения скрипта
        "active_operations": "scripts_only",  # Ключ: активные операции ("scripts_only", "json_only", "both")
        "domain": "rating.example.com",  # Ключ: домен для API запросов
        "params": {  # Ключ: параметры API запросов
            "api_endpoint": "/api/rating/participants",  # Ключ: конечная точка API
            "sort_by": "total_points",  # Ключ: поле для сортировки
            "include_rewards": True,  # Ключ: включать ли награды
            "include_crystals": True,  # Ключ: включать ли кристаллы
            "limit": 500  # Ключ: лимит записей
        },
        "data_source": "file",  # Ключ: источник данных
        "input_format": "CSV",  # Ключ: формат входного файла
        "csv_column": "participant_id",  # Ключ: название столбца для извлечения данных
        "csv_delimiter": ";",  # Ключ: разделитель в CSV файле
        "csv_encoding": "utf-8",  # Ключ: кодировка CSV файла
        "json_file": "rating_list"  # Ключ: имя JSON файла для обработки (без расширения)
    }
}



# =============================================================================
# ГЛОБАЛЬНЫЕ ПЕРЕМЕННЫЕ
# =============================================================================

logger = None  # Глобальный объект логгера (инициализируется в setup_logging())
program_start_time = None  # Время начала выполнения программы (записывается в main())
function_execution_times = {}  # Словарь для хранения времени выполнения функций (заполняется декоратором measure_time)
processed_actions_count = 0  # Счетчик обработанных действий (увеличивается в процессе работы программы)

# =============================================================================
# НАСТРОЙКА ЛОГИРОВАНИЯ
# =============================================================================

def setup_logging():
    """
    Настройка системы логирования
    
    Создает логгер с двумя обработчиками:
    - FileHandler: записывает логи в файл
    - StreamHandler: выводит логи в консоль
    
    Returns:
        logging.Logger: Настроенный объект логгера
    """
    global logger
    
    # Создание директорий если не существуют
    # Используем exist_ok=True чтобы не вызывать ошибку если директория уже существует
    log_dir = os.path.join(BASE_DIR, SUBDIRECTORIES["LOGS"])
    input_dir = os.path.join(BASE_DIR, SUBDIRECTORIES["INPUT"])
    output_dir = os.path.join(BASE_DIR, SUBDIRECTORIES["OUTPUT"])
    script_dir = os.path.join(BASE_DIR, SUBDIRECTORIES["SCRIPT"])
    config_dir = os.path.join(BASE_DIR, SUBDIRECTORIES["CONFIG"])
    json_dir = os.path.join(BASE_DIR, SUBDIRECTORIES["JSON"])
    
    os.makedirs(log_dir, exist_ok=True)
    os.makedirs(input_dir, exist_ok=True)
    os.makedirs(output_dir, exist_ok=True)
    os.makedirs(script_dir, exist_ok=True)
    os.makedirs(config_dir, exist_ok=True)
    os.makedirs(json_dir, exist_ok=True)
    
    # Формирование имени файла лога с временной меткой
    # Формат: LOG_DEBUG_2024-01-15.log
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d")
    log_filename = f"{LOG_FILENAME_BASE}_{LOG_LEVEL}_{timestamp}.log"
    log_filepath = os.path.join(log_dir, log_filename)
    
    # Настройка основного логгера
    logger = logging.getLogger('GameScriptGenerator')
    logger.setLevel(getattr(logging, LOG_LEVEL))
    
    # Удаление существующих handlers для избежания дублирования
    for handler in logger.handlers[:]:
        logger.removeHandler(handler)
    
    # Создание file handler для записи в файл
    file_handler = logging.FileHandler(log_filepath, encoding='utf-8')
    file_handler.setLevel(getattr(logging, LOG_LEVEL))
    
    # Создание console handler для вывода в консоль
    console_handler = logging.StreamHandler()
    console_handler.setLevel(getattr(logging, LOG_LEVEL))
    
    # Создание форматтера для логов
    # Включает время с миллисекундами, имя логгера, уровень и сообщение
    formatter = logging.Formatter(
        '%(asctime)s.%(msecs)03d - %(name)s - %(levelname)s - %(message)s',
        datefmt='%Y-%m-%d %H:%M:%S'
    )
    
    # Применение форматтера к обработчикам
    file_handler.setFormatter(formatter)
    console_handler.setFormatter(formatter)
    
    # Добавление обработчиков к логгеру
    logger.addHandler(file_handler)
    logger.addHandler(console_handler)
    
    return logger

def get_script_logger(script_name, stage=None):
    """
    Получение логгера для конкретного скрипта и стадии
    
    Args:
        script_name (str): Название скрипта (leaders_for_admin, reward, etc.)
        stage (str, optional): Стадия обработки (generation, processing, conversion)
        
    Returns:
        logging.Logger: Логгер с конкретным названием
    """
    if stage:
        logger_name = f'GameScriptGenerator.{script_name}.{stage}'
    else:
        logger_name = f'GameScriptGenerator.{script_name}'
    
    script_logger = logging.getLogger(logger_name)
    
    # Если у логгера нет обработчиков, наследуем от основного
    if not script_logger.handlers:
        # Убеждаемся, что основной логгер инициализирован
        if logger is None:
            setup_logging()
        
        if logger and logger.handlers:
            script_logger.handlers = logger.handlers.copy()
            script_logger.setLevel(logger.level)
            script_logger.propagate = False
        else:
            # Fallback: используем основной логгер
            script_logger = logger
    
    return script_logger

# =============================================================================
# ДЕКОРАТОРЫ ДЛЯ ИЗМЕРЕНИЯ ВРЕМЕНИ ВЫПОЛНЕНИЯ
# =============================================================================

def measure_time(func):
    """
    Декоратор для измерения времени выполнения функций
    
    Логирует начало и конец выполнения функции, а также время выполнения.
    Сохраняет время выполнения в глобальный словарь function_execution_times.
    
    Args:
        func: Функция для декорирования
        
    Returns:
        wrapper: Обернутая функция с измерением времени
    """
    @wraps(func)
    def wrapper(*args, **kwargs):
        start_time = time.time()
        
        # Формирование строки параметров для логирования
        # Ограничиваем количество аргументов для читаемости логов
        # Исключаем вывод содержимого скриптов
        if func.__name__ in ['generate_leaders_for_admin_script', 'generate_reward_script']:
            params_str = f"args=(), kwargs={list(kwargs.keys())}"
        else:
            params_str = f"args={args[:2] if len(args) > 2 else args}, kwargs={list(kwargs.keys())}"
        logger.debug(LOG_MESSAGES['function_start'].format(func=func.__name__, params=params_str))
        
        try:
            # Выполнение функции
            result = func(*args, **kwargs)
            execution_time = time.time() - start_time
            
            # Сохранение времени выполнения в глобальный словарь
            function_execution_times[func.__name__] = execution_time
            
            # Логирование успешного завершения
            # Исключаем вывод содержимого скриптов
            if func.__name__ in ['generate_leaders_for_admin_script', 'generate_reward_script']:
                logger.debug(LOG_MESSAGES['function_completed'].format(func=func.__name__, params="args=(), kwargs=[]", time=f"{execution_time:.4f}"))
            else:
                logger.debug(LOG_MESSAGES['function_completed'].format(func=func.__name__, params=params_str, time=f"{execution_time:.4f}"))
            return result
            
        except Exception as e:
            # Обработка ошибок
            execution_time = time.time() - start_time
            function_execution_times[func.__name__] = execution_time
            # Исключаем вывод содержимого скриптов при ошибках
            if func.__name__ in ['generate_leaders_for_admin_script', 'generate_reward_script']:
                logger.error(LOG_MESSAGES['function_error'].format(func=func.__name__, params="args=(), kwargs=[]", error=str(e)))
            else:
                logger.error(LOG_MESSAGES['function_error'].format(func=func.__name__, params=params_str, error=str(e)))
            raise
            
    return wrapper

# =============================================================================
# ФУНКЦИИ ОБРАБОТКИ ДАННЫХ
# =============================================================================

@measure_time
def load_data_from_file(filepath, file_format="TXT", csv_delimiter=None, csv_encoding=None, csv_column=None):
    """
    Загрузка данных из файла
    
    Поддерживает форматы TXT и CSV. Для TXT файлов использует массив разделителей,
    для CSV файлов - указанный разделитель и столбец.
    
    Args:
        filepath (str): Путь к файлу для загрузки
        file_format (str): Формат файла ("TXT" или "CSV")
        csv_delimiter (str): Разделитель для CSV файлов (по умолчанию из констант)
        csv_encoding (str): Кодировка для CSV файлов (по умолчанию из констант)
        csv_column (str): Название столбца для CSV файлов (по умолчанию из констант)
        
    Returns:
        list: Список загруженных данных
    """
    global processed_actions_count
    
    # Использование переданных параметров или значений по умолчанию
    delimiter = csv_delimiter or CSV_DELIMITER
    encoding = csv_encoding or CSV_ENCODING
    column = csv_column or CSV_COLUMN_NAME
    
    logger.debug(LOG_MESSAGES['file_loading'].format(file_path=filepath, format=file_format))
    
    # Проверка существования файла
    if not os.path.exists(filepath):
        logger.error(LOG_MESSAGES['file_not_found'].format(file_path=filepath))
        return []
    
    data_list = []
    
    try:
        if file_format.upper() == "TXT":
            # Обработка текстового файла
            with open(filepath, 'r', encoding='utf-8') as file:
                content = file.read()
                delimiters_found = 0
                
                # Разделение по массиву разделителей
                # Заменяем все найденные разделители на единый разделитель
                for delimiter_char in TXT_DELIMITERS:
                    if delimiter_char in content:
                        delimiters_found += 1
                        content = content.replace(delimiter_char, '|SPLIT|')
                
                logger.debug(LOG_MESSAGES['txt_processing'].format(delimiters_count=delimiters_found))
                
                # Разделяем по единому разделителю и очищаем пустые элементы
                data_list = [item.strip() for item in content.split('|SPLIT|') 
                           if item.strip() and item.strip() != '|SPLIT|']
                
        elif file_format.upper() == "CSV":
            # Обработка CSV файла
            logger.debug(LOG_MESSAGES['csv_processing'].format(delimiter=delimiter, encoding=encoding, column=column))
            with open(filepath, 'r', encoding=encoding) as file:
                csv_reader = csv.DictReader(file, delimiter=delimiter)
                for row in csv_reader:
                    # Извлекаем данные из указанного столбца
                    if column in row and row[column].strip():
                        data_list.append(row[column].strip())
                        
        # Обновление счетчика обработанных действий
        processed_actions_count += len(data_list)
        logger.info(LOG_MESSAGES['file_loaded'].format(file_path=filepath, count=len(data_list)))
        
    except Exception as e:
        logger.error(LOG_MESSAGES['file_load_error'].format(file_path=filepath, error=str(e)))
        
    return data_list

@measure_time 
def get_data():
    """
    Получение данных согласно настройкам
    
    В зависимости от значения DATA_SOURCE загружает данные из файла
    или возвращает тестовые данные из переменной.
    
    Returns:
        list: Список данных для обработки
    """
    # Эта функция теперь используется только для тестовых данных
    # Для каждого скрипта данные загружаются индивидуально в соответствующих функциях
    logger.info(LOG_MESSAGES['using_test_data'].format(count=len(TEST_DATA_LIST)))
    return TEST_DATA_LIST.copy()

@measure_time
def save_script_to_file(script_content, script_name, config_key=None, variant=None):
    """
    Сохранение сгенерированного скрипта в файл TXT
    
    Создает файл с именем на основе названия скрипта, варианта и временной метки
    в директории SCRIPT.
    
    Args:
        script_content (str): Содержимое скрипта для сохранения
        script_name (str): Название скрипта для формирования имени файла
        config_key (str, optional): Ключ конфигурации для дополнительной информации
        variant (str, optional): Вариант скрипта (sigma/alpha)
        
    Returns:
        str: Путь к сохраненному файлу или None в случае ошибки
    """
    try:
        # Создание временной метки
        timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        
        # Формирование имени файла
        # Убираем специальные символы и заменяем пробелы на подчеркивания
        safe_name = script_name.replace(' ', '_').replace('(', '').replace(')', '').replace('/', '_')
        
        # Добавляем информацию о варианте если есть
        if variant:
            filename = f"{safe_name}_{variant.upper()}_{timestamp}.txt"
        else:
            filename = f"{safe_name}_{timestamp}.txt"
        
        # Полный путь к файлу (используем папку SCRIPT)
        script_dir = os.path.join(BASE_DIR, SUBDIRECTORIES["SCRIPT"])
        filepath = os.path.join(script_dir, filename)
        
        # Создание директории если не существует
        os.makedirs(script_dir, exist_ok=True)
        
        # Сохранение скрипта в файл
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(script_content)
        
        logger.info(LOG_MESSAGES['script_saved'].format(file_path=filepath))
        return filepath
        
    except Exception as e:
        logger.error(LOG_MESSAGES['script_save_error'].format(error=str(e)))
        return None

# Функция копирования в буфер обмена удалена - не используется
# @measure_time
# def copy_to_clipboard(text):
#     """
#     Копирование текста в буфер обмена
#     
#     Использует библиотеку pyperclip для копирования текста в системный буфер обмена.
#     
#     Args:
#         text (str): Текст для копирования в буфер обмена
#         
#     Returns:
#         bool: True если копирование успешно, False в случае ошибки
#     """
#     try:
#         pyperclip.copy(text)
#         logger.debug(LOG_MESSAGES['clipboard_copied'])
#         return True
#     except Exception as e:
#         logger.error(LOG_MESSAGES['clipboard_error'].format(error=str(e)))
#         return False

# =============================================================================
# ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ ДЛЯ JSON И EXCEL
# =============================================================================

def parse_float_safe(val, context=None):
    """Безопасное преобразование в float с обработкой европейского формата"""
    if val is None or val == "":
        return None
    try:
        # Обработка европейского формата чисел (запятая вместо точки)
        if isinstance(val, str):
            # Удаляем пробелы и заменяем запятую на точку
            val = val.replace(' ', '').replace(',', '.')
            # Удаляем неразрывные пробелы и другие специальные символы
            val = val.replace('\u2009', '').replace('\u00a0', '')
        return float(val)
    except (ValueError, TypeError) as ex:
        if context:
            logger.warning(LOG_MESSAGES['float_conversion_error'].format(val=val, ex=ex, context=context))
        return None

def flatten_leader_data(leader_data):
    """Преобразование данных лидера в плоскую структуру"""
    flattened = {}
    
    # Основные поля из структуры LeadersForAdmin
    flattened['employeeNumber'] = leader_data.get('employeeNumber', '')
    flattened['lastName'] = leader_data.get('lastName', '')
    flattened['firstName'] = leader_data.get('firstName', '')
    flattened['indicatorValue'] = leader_data.get('indicatorValue', '')
    flattened['successValue'] = leader_data.get('successValue', '')
    flattened['terDivisionName'] = leader_data.get('terDivisionName', '')
    flattened['employeeStatus'] = leader_data.get('employeeStatus', '')
    flattened['businessBlock'] = leader_data.get('businessBlock', '')
    
    # Поля турнира (добавлены при обработке всех турниров)
    flattened['tournamentId'] = leader_data.get('tournamentId', '')
    flattened['tournamentIndicator'] = leader_data.get('tournamentIndicator', '')
    flattened['tournamentStatus'] = leader_data.get('tournamentStatus', '')
    flattened['contestants'] = leader_data.get('contestants', '')
    
    # Создаем полное имя
    flattened['fullName'] = f"{leader_data.get('lastName', '')} {leader_data.get('firstName', '')}".strip()
    
    # Обработка вложенной структуры divisionRatings
    division_ratings = leader_data.get('divisionRatings', [])
    
    # Инициализируем колонки для каждой категории (BANK, TB, GOSB)
    categories = ['BANK', 'TB', 'GOSB']
    for category in categories:
        flattened[f'{category}_groupId'] = ''
        flattened[f'{category}_placeInRating'] = ''
        flattened[f'{category}_ratingCategoryName'] = ''
    
    # Заполняем данные из divisionRatings
    for rating in division_ratings:
        group_code = rating.get('groupCode', '')
        if group_code in categories:
            # Преобразуем groupId в целое число
            group_id_raw = rating.get('groupId', '')
            try:
                flattened[f'{group_code}_groupId'] = int(float(group_id_raw)) if group_id_raw else ''
            except (ValueError, TypeError):
                flattened[f'{group_code}_groupId'] = ''
            
            # Преобразуем placeInRating в целое число
            place_in_rating_raw = rating.get('placeInRating', '')
            try:
                flattened[f'{group_code}_placeInRating'] = int(float(place_in_rating_raw)) if place_in_rating_raw else ''
            except (ValueError, TypeError):
                flattened[f'{group_code}_placeInRating'] = ''
            
            flattened[f'{group_code}_ratingCategoryName'] = rating.get('ratingCategoryName', '')
    
    return flattened

def flatten_reward_profile_data(profile_data):
    """
    Преобразование данных профиля награды в плоскую структуру
    
    Args:
        profile_data (dict): Данные профиля из API наград
        
    Returns:
        dict: Плоская структура данных профиля
    """
    flattened = {}
    
    # Основные поля профиля
    flattened['rewardCode'] = profile_data.get('rewardCode', '')
    flattened['badgeName'] = profile_data.get('badgeName', '')
    flattened['badgeDescription'] = profile_data.get('badgeDescription', '')
    flattened['badgeType'] = profile_data.get('badgeType', '')
    flattened['badgeCategory'] = profile_data.get('badgeCategory', '')
    
    # Поля профиля сотрудника
    flattened['employeeNumber'] = profile_data.get('employeeNumber', '')
    flattened['lastName'] = profile_data.get('lastName', '')
    flattened['firstName'] = profile_data.get('firstName', '')
    flattened['middleName'] = profile_data.get('middleName', '')
    flattened['fullName'] = profile_data.get('fullName', '')
    
    # Контактная информация
    flattened['email'] = profile_data.get('email', '')
    flattened['phone'] = profile_data.get('phone', '')
    flattened['mobilePhone'] = profile_data.get('mobilePhone', '')
    
    # Организационная информация
    flattened['terDivisionName'] = profile_data.get('terDivisionName', '')
    flattened['divisionName'] = profile_data.get('divisionName', '')
    flattened['departmentName'] = profile_data.get('departmentName', '')
    flattened['positionName'] = profile_data.get('positionName', '')
    flattened['employeeStatus'] = profile_data.get('employeeStatus', '')
    flattened['businessBlock'] = profile_data.get('businessBlock', '')
    
    # Информация о награде
    flattened['awardDate'] = profile_data.get('awardDate', '')
    flattened['awardReason'] = profile_data.get('awardReason', '')
    flattened['awardLevel'] = profile_data.get('awardLevel', '')
    flattened['awardValue'] = profile_data.get('awardValue', '')
    
    # Статистика и показатели
    flattened['indicatorValue'] = profile_data.get('indicatorValue', '')
    flattened['successValue'] = profile_data.get('successValue', '')
    flattened['rating'] = profile_data.get('rating', '')
    flattened['placeInRating'] = profile_data.get('placeInRating', '')
    
    # Дополнительные поля
    flattened['photoUrl'] = profile_data.get('photoUrl', '')
    flattened['isActive'] = profile_data.get('isActive', '')
    flattened['lastActivityDate'] = profile_data.get('lastActivityDate', '')
    
    # Создаем полное имя, если его нет
    if not flattened['fullName']:
        name_parts = [flattened['lastName'], flattened['firstName'], flattened['middleName']]
        flattened['fullName'] = ' '.join([part for part in name_parts if part]).strip()
    
    # Обработка вложенных структур (если есть)
    if 'divisionRatings' in profile_data:
        division_ratings = profile_data['divisionRatings']
        categories = ['BANK', 'TB', 'GOSB']
        for category in categories:
            flattened[f'{category}_groupId'] = ''
            flattened[f'{category}_placeInRating'] = ''
            flattened[f'{category}_ratingCategoryName'] = ''
        
        for rating in division_ratings:
            if isinstance(rating, dict):
                group_code = rating.get('groupCode', '')
                if group_code in categories:
                    group_id_raw = rating.get('groupId', '')
                    try:
                        flattened[f'{group_code}_groupId'] = int(float(group_id_raw)) if group_id_raw else ''
                    except (ValueError, TypeError):
                        flattened[f'{group_code}_groupId'] = ''
                    
                    place_in_rating_raw = rating.get('placeInRating', '')
                    try:
                        flattened[f'{group_code}_placeInRating'] = int(float(place_in_rating_raw)) if place_in_rating_raw else ''
                    except (ValueError, TypeError):
                        flattened[f'{group_code}_placeInRating'] = ''
                    
                    flattened[f'{group_code}_ratingCategoryName'] = rating.get('ratingCategoryName', '')
    
    return flattened

def flatten_reward_leader_data(leader_data, reward_code):
    """
    Преобразование данных лидера награды в плоскую структуру
    
    Args:
        leader_data (dict): Данные лидера из структуры наград
        reward_code (str): Код награды
        
    Returns:
        dict: Плоская структура данных лидера награды
    """
    flattened = {}
    
    # Основные поля лидера
    flattened['rewardCode'] = reward_code
    flattened['employeeNumber'] = leader_data.get('employeeNumber', '')
    flattened['lastName'] = leader_data.get('lastName', '')
    flattened['firstName'] = leader_data.get('firstName', '')
    flattened['terDivisionName'] = leader_data.get('terDivisionName', '')
    flattened['gosbCode'] = leader_data.get('gosbCode', '')
    flattened['employeeStatus'] = leader_data.get('employeeStatus', '')
    flattened['receivingDate'] = leader_data.get('receivingDate', '')
    flattened['isMarked'] = leader_data.get('isMarked', False)
    
    # Создаем полное имя
    flattened['fullName'] = f"{leader_data.get('lastName', '')} {leader_data.get('firstName', '')}".strip()
    
    # Обработка colorCode
    color_code = leader_data.get('colorCode', {})
    flattened['colorPrimary'] = color_code.get('primary', '')
    flattened['colorSecondary'] = color_code.get('secondary', '')
    
    # Обработка earnedBadges
    earned_badges = leader_data.get('earnedBadges', [])
    flattened['earnedBadgesCount'] = len(earned_badges)
    flattened['earnedBadgesList'] = ', '.join([badge.get('name', '') for badge in earned_badges if badge.get('name')])
    
    # Обработка tags
    tags = leader_data.get('tags', [])
    flattened['tagsCount'] = len(tags)
    flattened['tagsList'] = ', '.join([tag.get('tagName', '') for tag in tags if tag.get('tagName')])
    
    # Детальная информация о тегах
    for i, tag in enumerate(tags[:5]):  # Ограничиваем до 5 тегов
        flattened[f'tag{i+1}_id'] = tag.get('tagId', '')
        flattened[f'tag{i+1}_name'] = tag.get('tagName', '')
        flattened[f'tag{i+1}_color'] = tag.get('tagColor', '')
    
    # Заполняем пустые теги
    for i in range(len(tags), 5):
        flattened[f'tag{i+1}_id'] = ''
        flattened[f'tag{i+1}_name'] = ''
        flattened[f'tag{i+1}_color'] = ''
    
    return flattened

def apply_excel_styling(workbook, freeze_cell="B2"):
    """Применение стилей к Excel файлу"""
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        
        # Стили для заголовков
        header_fill = PatternFill(start_color=EXCEL_COLORS["header"], end_color=EXCEL_COLORS["header"], fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Стили для подзаголовков
        subheader_fill = PatternFill(start_color=EXCEL_COLORS["subheader"], end_color=EXCEL_COLORS["subheader"], fill_type="solid")
        subheader_font = Font(bold=True)
        
        # Применение стилей к заголовкам
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Закрепление строк и столбцов (если есть данные)
        if worksheet.max_row > 1:
            worksheet.freeze_panes = freeze_cell
        
        # Автофильтр для листа DATA
        if sheet_name == 'DATA' and worksheet.max_row > 1:
            # Получаем диапазон данных для автофильтра
            max_col = worksheet.max_column
            max_row = worksheet.max_row
            filter_range = f"A1:{get_column_letter(max_col)}{max_row}"
            worksheet.auto_filter.ref = filter_range
        
        # Автоматическая ширина столбцов
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

def create_summary_sheet(workbook, data_df):
    """Создание листа с сводной информацией"""
    if 'DATA' not in workbook.sheetnames:
        return
    
    # Создаем лист SUMMARY
    if 'SUMMARY' in workbook.sheetnames:
        workbook.remove(workbook['SUMMARY'])
    summary_sheet = workbook.create_sheet('SUMMARY')
    
    # Основная статистика
    summary_data = [
        ['Параметр', 'Значение'],
        ['Общее количество участников', len(data_df)],
        ['Участники с номером сотрудника', len(data_df[data_df['employeeNumber'].notna() & (data_df['employeeNumber'] != '')])],
        ['Участники со статусом CONTESTANT', len(data_df[data_df['employeeStatus'] == 'CONTESTANT'])],
        ['Среднее значение показателя', round(data_df['indicatorValue_numeric'].mean(), 2) if 'indicatorValue_numeric' in data_df.columns else 'N/A'],
        ['Максимальное значение показателя', data_df['indicatorValue_numeric'].max() if 'indicatorValue_numeric' in data_df.columns else 'N/A'],
        ['Минимальное значение показателя', data_df['indicatorValue_numeric'].min() if 'indicatorValue_numeric' in data_df.columns else 'N/A'],
    ]
    
    # Добавляем данные в лист
    for row_idx, row_data in enumerate(summary_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            summary_sheet.cell(row=row_idx, column=col_idx, value=value)
    
    # Применяем стили
    header_fill = PatternFill(start_color=EXCEL_COLORS["header"], end_color=EXCEL_COLORS["header"], fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for cell in summary_sheet[1]:
        cell.fill = header_fill
        cell.font = header_font

def create_statistics_sheet(workbook, data_df):
    """Создание листа со статистикой"""
    if 'DATA' not in workbook.sheetnames:
        return
    
    # Создаем лист STATISTICS
    if 'STATISTICS' in workbook.sheetnames:
        workbook.remove(workbook['STATISTICS'])
    stats_sheet = workbook.create_sheet('STATISTICS')
    
    # Статистика по департаментам
    if 'terDivisionName' in data_df.columns:
        dept_stats = data_df['terDivisionName'].value_counts().reset_index()
        dept_stats.columns = ['Территориальное подразделение', 'Количество участников']
        
        # Добавляем заголовок
        stats_sheet.cell(row=1, column=1, value='Статистика по территориальным подразделениям')
        stats_sheet.cell(row=1, column=1).font = Font(bold=True, size=14)
        
        # Добавляем данные
        for row_idx, (_, row_data) in enumerate(dept_stats.iterrows(), 3):
            stats_sheet.cell(row=row_idx, column=1, value=row_data['Территориальное подразделение'])
            stats_sheet.cell(row=row_idx, column=2, value=row_data['Количество участников'])
    
    # Статистика по бизнес-блокам
    if 'businessBlock' in data_df.columns:
        block_stats = data_df['businessBlock'].value_counts().reset_index()
        block_stats.columns = ['Бизнес-блок', 'Количество участников']
        
        # Добавляем заголовок
        stats_sheet.cell(row=1, column=4, value='Статистика по бизнес-блокам')
        stats_sheet.cell(row=1, column=4).font = Font(bold=True, size=14)
        
        # Добавляем данные
        for row_idx, (_, row_data) in enumerate(block_stats.iterrows(), 3):
            stats_sheet.cell(row=row_idx, column=4, value=row_data['Бизнес-блок'])
            stats_sheet.cell(row=row_idx, column=5, value=row_data['Количество участников'])
    
    # Применяем стили
    header_fill = PatternFill(start_color=EXCEL_COLORS["subheader"], end_color=EXCEL_COLORS["subheader"], fill_type="solid")
    header_font = Font(bold=True)
    
    for cell in stats_sheet[3]:
        cell.fill = header_fill
        cell.font = header_font

def create_reward_summary_sheet(workbook, data_df):
    """Создание сводного листа для данных наград"""
    if 'DATA' not in workbook.sheetnames:
        return
    
    # Создаем лист REWARD_SUMMARY
    if 'REWARD_SUMMARY' in workbook.sheetnames:
        workbook.remove(workbook['REWARD_SUMMARY'])
    summary_sheet = workbook.create_sheet('REWARD_SUMMARY')
    
    # Сводная статистика по наградам
    summary_data = [
        ["Параметр", "Значение"],
        ["Общее количество профилей с наградами", len(data_df)],
        ["Количество уникальных кодов наград", data_df['rewardCode'].nunique() if 'rewardCode' in data_df.columns else 0],
        ["Количество уникальных сотрудников", data_df['employeeNumber'].nunique() if 'employeeNumber' in data_df.columns else 0],
        ["Количество уникальных подразделений", data_df['terDivisionName'].nunique() if 'terDivisionName' in data_df.columns else 0]
    ]
    
    # Статистика по типам наград
    if 'badgeType' in data_df.columns:
        badge_type_stats = data_df['badgeType'].value_counts()
        summary_data.append(["", ""])
        summary_data.append(["Статистика по типам наград", ""])
        for badge_type, count in badge_type_stats.items():
            summary_data.append([badge_type, count])
    
    # Статистика по категориям наград
    if 'badgeCategory' in data_df.columns:
        badge_category_stats = data_df['badgeCategory'].value_counts()
        summary_data.append(["", ""])
        summary_data.append(["Статистика по категориям наград", ""])
        for badge_category, count in badge_category_stats.items():
            summary_data.append([badge_category, count])
    
    # Статистика по структурам данных
    if 'structure' in data_df.columns:
        structure_stats = data_df['structure'].value_counts()
        summary_data.append(["", ""])
        summary_data.append(["Статистика по структурам данных", ""])
        for structure, count in structure_stats.items():
            summary_data.append([structure, count])
    
    # Записываем данные в лист
    for row_idx, row_data in enumerate(summary_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            summary_sheet.cell(row=row_idx, column=col_idx, value=value)
    
    # Применяем стили
    header_fill = PatternFill(start_color=EXCEL_COLORS["subheader"], end_color=EXCEL_COLORS["subheader"], fill_type="solid")
    header_font = Font(bold=True)
    
    for cell in summary_sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
    
    logger.info(LOG_MESSAGES['reward_summary_sheet_created'])

# =============================================================================
# ФУНКЦИИ ГЕНЕРАЦИИ СКРИПТОВ
# =============================================================================

def load_script_data(config_key, data_list=None):
    """
    Общая функция для загрузки данных для скриптов
    
    Args:
        config_key (str): Ключ конфигурации из FUNCTION_CONFIGS
        data_list (list, optional): Список данных для обработки
        
    Returns:
        tuple: (config, data_list, variants_configs)
    """
    config = FUNCTION_CONFIGS[config_key]
    
    # Получение данных согласно конфигурации
    if data_list is None:
        if config["data_source"] == "file":
            # Загрузка данных из файла
            file_extension = FILE_EXTENSIONS.get(config["input_format"], ".txt")
            filename = f"{config_key}_data{file_extension}"
            config_dir = os.path.join(BASE_DIR, SUBDIRECTORIES["CONFIG"])
            filepath = os.path.join(config_dir, filename)
            data_list = load_data_from_file(
                filepath, 
                config["input_format"],
                config["csv_delimiter"],
                config["csv_encoding"],
                config["csv_column"]
            )
        elif config["data_source"] == "external_file":
            # Загрузка данных из внешнего файла
            file_extension = FILE_EXTENSIONS.get(config["input_format"], ".csv")
            config_dir = os.path.join(BASE_DIR, SUBDIRECTORIES["CONFIG"])
            filepath = os.path.join(config_dir, config["input_file"] + file_extension)
            data_list = load_data_from_file(
                filepath, 
                config["input_format"],
                config["csv_delimiter"],
                config["csv_encoding"],
                config["csv_column"]
            )
        else:
            # Использование тестовых данных
            data_list = TEST_DATA_LIST.copy()
    
    # Получение всех вариантов конфигурации
    variants_configs = config["variants"]
    
    # Логирование процесса генерации
    logger.debug(LOG_MESSAGES['script_generation'].format(script_name=f"{config['name']} (ALL VARIANTS)"))
    logger.debug(LOG_MESSAGES['config_loaded'].format(script_name=f"{config['name']} (ALL VARIANTS)"))
    logger.debug(f"Варианты: {', '.join([v.upper() for v in variants_configs.keys()])}")
    logger.debug(LOG_MESSAGES['data_source_selected'].format(
        source=config['data_source'], 
        format=config['input_format']
    ))
    
    return config, data_list, variants_configs

def save_and_copy_script(script, config, config_key, data_list):
    """
    Общая функция для сохранения скрипта
    Args:
        script (str): Сгенерированный JavaScript скрипт
        config (dict): Конфигурация скрипта
        config_key (str): Ключ конфигурации
        data_list (list): Список данных
    """
    # Сохранение скрипта в файл
    saved_filepath = save_script_to_file(script, config['name'], config_key)
    logger.info(LOG_MESSAGES['script_generated'].format(script_name=config['name'], count=len(data_list)))

# =============================================================================
# ФУНКЦИИ ГЕНЕРАЦИИ JAVASCRIPT СКРИПТОВ (ЗАГЛУШКИ)
# =============================================================================

@measure_time
def generate_leaders_for_admin_script(data_list=None):
    """
    Генерация скрипта для получения информации по участникам турнира
    
    Args:
        data_list (list, optional): Список ID участников
        
    Returns:
        str: Сгенерированный JavaScript скрипт
    """
    script_logger = get_script_logger("leaders_for_admin", "generation")
    script_logger.info(LOG_MESSAGES['script_generation_start'].format(script_name="LeadersForAdmin"))
    script_logger.debug(LOG_MESSAGES['function_start'].format(func="generate_leaders_for_admin_script", params=f"args=({data_list}), kwargs=[]"))
    
    # Загрузка данных и конфигурации
    script_logger.info(LOG_MESSAGES['data_loading'])
    config, data_list, variants_configs = load_script_data("leaders_for_admin", data_list)
    
    script_logger.info(LOG_MESSAGES['config_loaded_count'].format(count=len(data_list)))
    script_logger.debug(f"Варианты: {', '.join([v.upper() for v in variants_configs.keys()])}")
    
    # Получаем настройки из конфигурации
    delay = config.get('delay_between_requests', 5)
    max_retries = config.get('retry_count', 3)
    timeout = config.get('timeout', 30000)
    remove_photo_data = config.get('processing_options', {}).get('remove_photo_data', True)
    
    script_logger.debug(LOG_MESSAGES['request_params'].format(delay=delay, max_retries=max_retries, timeout=timeout))
    script_logger.debug(f"Удаление photoData: {remove_photo_data}")
    
    # Генерируем скрипты для всех вариантов
    generated_scripts = []
    
    for variant_name, variant_config in variants_configs.items():
        script_logger.info(f"Генерация скрипта для варианта: {variant_name.upper()}")
        script_logger.debug(LOG_MESSAGES['domain_info'].format(domain=variant_config['domain']))
        script_logger.debug(LOG_MESSAGES['api_path_info'].format(api_path=variant_config['params']['api_path']))
        
        # Генерация JavaScript скрипта для LeadersForAdmin
        script = f"""// ==UserScript==
// Скрипт для DevTools. Выгрузка лидеров для всех Tournament ID (одна страница на турнир)
// Вариант: {variant_name.upper()}
// Сгенерировано: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
// Количество турниров: {len(data_list)}
(async () => {{
  // === Удаление photoData рекурсивно ===
  function removePhotoData(obj) {{
    if (Array.isArray(obj)) {{
      obj.forEach(removePhotoData);
    }} else if (obj && typeof obj === 'object') {{
      Object.keys(obj).forEach(key => {{
        if (key === 'photoData') {{
          delete obj[key];
        }} else {{
          removePhotoData(obj[key]);
        }}
      }});
    }}
  }}

  // === Генерация timestamp ===
  function getTimestamp() {{
    const d = new Date();
    const pad = n => n.toString().padStart(2, '0');
    return d.getFullYear().toString()
      + pad(d.getMonth() + 1)
      + pad(d.getDate())
      + '-' + pad(d.getHours())
      + pad(d.getMinutes())
      + pad(d.getSeconds());
  }}

  const ids = [{', '.join([f'"{item}"' for item in data_list])}];
  const service = 'leadersForAdmin';
  const BASE_URL = '{variant_config['domain']}{variant_config['params']['api_path']}';
  const results = {{}};
  let processed = 0, skipped = 0, errors = 0;
  console.log('▶️ Всего к обработке:', ids.length, 'код(ов)');
  console.log('🎯 Вариант:', '{variant_name.upper()}');

  for (let i = 0; i < ids.length; ++i) {{
    const tid = ids[i];
    const url = BASE_URL + tid + '/' + service + '?pageNum=1';
    console.log(`⏳ [${{i+1}}/${{ids.length}}] Обрабатываем код: ${{tid}}`);
    let resp, data;
    try {{
      resp = await fetch(url, {{
        headers: {{ 'Accept': 'application/json', 'Cookie': document.cookie }}, credentials: 'include'
      }});
      if (!resp.ok) {{
        console.warn(`❌ [${{i+1}}/${{ids.length}}] Код ${{tid}}: HTTP статус ${{resp.status}}`);
        errors++;
        continue;
      }}
      data = await resp.json();
      // Число участников
      let leadersCount = 0;
      try {{
        const leadersArr = data?.body?.tournament?.leaders || data?.body?.badge?.leaders;
        if (Array.isArray(leadersArr)) {{
          leadersCount = leadersArr.length;
        }}
      }} catch {{}}
      if (leadersCount === 0) {{
        console.log(`ℹ️ [${{i+1}}/${{ids.length}}] Код ${{tid}} пропущен: участников = 0`);
        skipped++;
        continue;
      }}
      console.log(`✅ [${{i+1}}/${{ids.length}}] Код ${{tid}}: успешно, участников: ${{leadersCount}}`);
      results[tid] = [data];
      processed++;
      await new Promise(r => setTimeout(r, {delay} * 1000));
    }} catch (e) {{
      console.error(`❌ [${{i+1}}/${{ids.length}}] Код ${{tid}}: Ошибка запроса:`, e);
      errors++;
    }}
  }}

  // Удаляем photoData только если это включено в настройках
  if ({str(remove_photo_data).lower()}) {{
    console.log('🧹 Удаляем все поля photoData');
    removePhotoData(results);
  }} else {{
    console.log('🧹 Удаление photoData отключено в настройках');
  }}

  console.log('💾 Сохраняем файл ...');
  const ts = getTimestamp();
  const blob = new Blob([JSON.stringify(results, null, 2)], {{type: 'application/json'}});
  const a = document.createElement('a');
  a.href = URL.createObjectURL(blob);
  a.download = service + '_{variant_name.upper()}_' + ts + '.json';
  document.body.appendChild(a);
  a.click();
  a.remove();
  console.log(`🏁 Обработка завершена. Всего: ${{ids.length}}. Успешно: ${{processed}}. Пропущено: ${{skipped}}. Ошибок: ${{errors}}. Файл скачан.`);
}})();"""
        
        # Сохранение скрипта для текущего варианта
        script_logger.info(f"Сохранение скрипта для варианта: {variant_name.upper()}")
        saved_filepath = save_script_to_file(script, config['name'], "leaders_for_admin", variant_name)
        generated_scripts.append((variant_name, saved_filepath))
    
    # Логирование результатов
    script_logger.info(f"Сгенерировано скриптов: {len(generated_scripts)}")
    for variant_name, filepath in generated_scripts:
        script_logger.info(f"✅ {variant_name.upper()}: {filepath}")
    
    script_logger.info(LOG_MESSAGES['script_generated_success'].format(script_name="LeadersForAdmin", count=len(data_list)))
    script_logger.debug(LOG_MESSAGES['function_completed'].format(func="generate_leaders_for_admin_script", params="args=(), kwargs=[]", time="0.0000"))
    
    # Возвращаем информацию о сгенерированных скриптах
    return generated_scripts

@measure_time
def generate_reward_script(data_list=None):
    """
    Генерация скрипта для выгрузки профилей участников по кодам наград с пагинацией
    Args:
        data_list (list, optional): Список кодов наград
    Returns:
        str: Сгенерированный JavaScript скрипт
    """
    script_logger = get_script_logger("reward", "generation")
    script_logger.info(LOG_MESSAGES['script_generation_start'].format(script_name="Reward"))
    script_logger.debug(LOG_MESSAGES['function_start'].format(func="generate_reward_script", params=f"args=({data_list}), kwargs=[]"))
    
    import datetime
    import json
    
    script_logger.info(LOG_MESSAGES['data_loading'])
    config, data_list, variants_configs = load_script_data("reward", data_list)
    
    script_logger.info(LOG_MESSAGES['config_loaded_count'].format(count=len(data_list)))
    script_logger.debug(f"Варианты: {', '.join([v.upper() for v in variants_configs.keys()])}")
    
    # Получаем настройки из конфигурации
    delay = config.get('delay_between_requests', 5)
    max_retries = config.get('retry_count', 3)
    timeout = config.get('timeout', 30000)
    remove_photo_data = config.get('processing_options', {}).get('remove_photo_data', True)
    
    script_logger.debug(LOG_MESSAGES['request_params'].format(delay=delay, max_retries=max_retries, timeout=timeout))
    script_logger.debug(f"Удаление photoData: {remove_photo_data}")
    
    # Генерируем скрипты для всех вариантов
    generated_scripts = []
    
    for variant_name, variant_config in variants_configs.items():
        script_logger.info(f"Генерация скрипта для варианта: {variant_name.upper()}")
        script_logger.debug(LOG_MESSAGES['domain_info'].format(domain=variant_config['domain']))
        script_logger.debug(LOG_MESSAGES['api_path_info'].format(api_path=variant_config['params']['api_path']))
        
        domain = variant_config['domain']
        api_path = variant_config['params']['api_path']
        service = variant_config['params']['service']
        base_url = f"{domain}{api_path}"
        
        script_logger.debug(LOG_MESSAGES['base_url_info'].format(base_url=base_url))
        
        ids_string = ', '.join([f'"{item}"' for item in data_list])
        script_logger.debug(LOG_MESSAGES['ids_generated'].format(count=len(data_list)))
        script = f'''// ==UserScript==
// Скрипт для DevTools. Выгрузка профилей участников по кодам наград с пагинацией
// Вариант: {variant_name.upper()}
(async () => {{
  function removePhotoData(obj) {{
    if (Array.isArray(obj)) {{ obj.forEach(removePhotoData); }}
    else if (obj && typeof obj === 'object') {{
      Object.keys(obj).forEach(key => {{
        if (key === 'photoData') delete obj[key];
        else removePhotoData(obj[key]);
      }});
    }}
  }}

  function getTimestamp() {{
    const d = new Date();
    const pad = n => n.toString().padStart(2, '0');
    return d.getFullYear().toString() + pad(d.getMonth() + 1) + pad(d.getDate()) + '-' + pad(d.getHours()) + pad(d.getMinutes()) + pad(d.getSeconds());
  }}

  function extractProfiles(data) {{
    try {{
      if (data?.body?.badge?.profiles && Array.isArray(data.body.badge.profiles)) {{
        return {{ profiles: data.body.badge.profiles }};
      }} else if (data?.body?.profiles && Array.isArray(data.body.profiles)) {{
        return {{ profiles: data.body.profiles }};
      }} else if (Array.isArray(data?.body)) {{
        return {{ profiles: data.body }};
      }} else if (Array.isArray(data)) {{
        return {{ profiles: data }};
      }}
      return null;
    }} catch (e) {{
      console.error('Ошибка при извлечении профилей:', e);
      return null;
    }}
  }}

  function extractContestantsCount(text) {{
    const match = text?.match(/(\\d+)/);
    return match ? parseInt(match[1], 10) : 0;
  }}

  async function fetchWithRetry(url, options, maxRetries = {max_retries}, timeout = {timeout}) {{
    for (let attempt = 1; attempt <= maxRetries; attempt++) {{
      try {{
        const controller = new AbortController();
        const id = setTimeout(() => controller.abort(), timeout);
        const response = await fetch(url, {{ ...options, signal: controller.signal }});
        clearTimeout(id);
        return response;
      }} catch (e) {{
        if (attempt === maxRetries) throw e;
        await new Promise(r => setTimeout(r, 1000 * attempt));
      }}
    }}
  }}

  const ids = [{ids_string}];
  const BASE_URL = '{base_url}';
  const results = {{}};
  let totalProfiles = 0;

  for (let i = 0; i < ids.length; i++) {{
    const code = ids[i];
    const baseUrl = `${{BASE_URL}}${{code}}/profiles`;
    console.log(`\\n🔍 [${{i + 1}}/${{ids.length}}] Код: ${{code}}`);
    
    try {{
      // Первый запрос для получения информации о количестве участников
      console.log(`📄 [${{i + 1}}/${{ids.length}}] Код: ${{code}} - Запрос страницы 1`);
      const firstResp = await fetchWithRetry(`${{baseUrl}}?pageNum=1&divisionLevel=BANK`, {{
        headers: {{ 'Accept': 'application/json', 'Cookie': document.cookie, 'User-Agent': navigator.userAgent }},
        credentials: 'include'
      }});
      
      if (!firstResp.ok) {{
        console.error(`❌ [${{i + 1}}/${{ids.length}}] Код: ${{code}} - HTTP ошибка: ${{firstResp.status}}`);
        continue;
      }}
      
      const firstData = await firstResp.json();
      const contestantsText = firstData?.body?.badge?.contestants;
      const count = extractContestantsCount(contestantsText);
      
      console.log(`👥 [${{i + 1}}/${{ids.length}}] Код: ${{code}} - Участников: ${{count}} (из текста: "${{contestantsText}}")`);
      
      if (count === 0) {{
        console.log(`⏭️ [${{i + 1}}/${{ids.length}}] Код: ${{code}} - Пропускаем (нет участников)`);
        continue;
      }}
      
      // Вычисляем количество страниц (делим на 100 с округлением вверх)
      const pagesCount = Math.ceil(count / 100);
      console.log(`📊 [${{i + 1}}/${{ids.length}}] Код: ${{code}} - Страниц для запроса: ${{pagesCount}} (участников: ${{count}}, по 100 на страницу)`);
      
      // Сохраняем первый запрос
      results[code] = [firstData];
      totalProfiles += (firstData?.body?.badge?.profiles?.length || 0);
      
      // Запрашиваем дополнительные страницы, если нужно
      if (pagesCount > 1) {{
        for (let page = 2; page <= pagesCount; page++) {{
          try {{
            console.log(`📄 [${{i + 1}}/${{ids.length}}] Код: ${{code}} - Запрос страницы ${{page}}/${{pagesCount}}`);
            const pageResp = await fetchWithRetry(`${{baseUrl}}?pageNum=${{page}}&divisionLevel=BANK`, {{
              headers: {{ 'Accept': 'application/json', 'Cookie': document.cookie, 'User-Agent': navigator.userAgent }},
              credentials: 'include'
            }});
            
            if (!pageResp.ok) {{
              console.error(`❌ [${{i + 1}}/${{ids.length}}] Код: ${{code}} - Страница ${{page}} - HTTP ошибка: ${{pageResp.status}}`);
              continue;
            }}
            
            const pageData = await pageResp.json();
            results[code].push(pageData);
            totalProfiles += (pageData?.body?.badge?.profiles?.length || 0);
            console.log(`✅ [${{i + 1}}/${{ids.length}}] Код: ${{code}} - Страница ${{page}}/${{pagesCount}} - Успешно`);
            
            // Задержка между запросами страниц
            if (page < pagesCount) {{
              await new Promise(resolve => setTimeout(resolve, {delay} * 1000));
            }}
          }} catch (pageError) {{
            console.error(`❌ [${{i + 1}}/${{ids.length}}] Код: ${{code}} - Страница ${{page}} - Ошибка:`, pageError);
          }}
        }}
      }}
      
      console.log(`✅ [${{i + 1}}/${{ids.length}}] Код: ${{code}} - Завершен, всего страниц: ${{results[code].length}}`);
      
    }} catch (e) {{
      console.error(`❌ [${{i + 1}}/${{ids.length}}] Код: ${{code}} - Критическая ошибка:`, e);
    }}
    
    // Задержка между кодами
    if (i < ids.length - 1) {{
      await new Promise(resolve => setTimeout(resolve, {delay} * 1000));
    }}
  }}

  // Удаляем photoData только если это включено в настройках
  if ({str(remove_photo_data).lower()}) {{
    console.log('\\n📦 Удаляем photoData...');
    removePhotoData(results);
  }} else {{
    console.log('\\n📦 Удаление photoData отключено в настройках');
  }}
  const ts = getTimestamp();
  const blob = new Blob([JSON.stringify(results, null, 2)], {{ type: 'application/json' }});
  const a = document.createElement('a');
  a.href = URL.createObjectURL(blob);
  a.download = `profiles_{variant_name.upper()}_${{ts}}.json`;
  a.click();
  console.log(`\\n✅ Завершено. Всего профилей: ${{totalProfiles}}`);
}})();
'''
        
        # Сохранение скрипта для текущего варианта
        script_logger.info(f"Сохранение скрипта для варианта: {variant_name.upper()}")
        saved_filepath = save_script_to_file(script, config['name'], "reward", variant_name)
        generated_scripts.append((variant_name, saved_filepath))
    
    # Логирование результатов
    script_logger.info(f"Сгенерировано скриптов: {len(generated_scripts)}")
    for variant_name, filepath in generated_scripts:
        script_logger.info(f"✅ {variant_name.upper()}: {filepath}")
    
    script_logger.info(LOG_MESSAGES['script_generated_success'].format(script_name="Reward", count=len(data_list)))
    script_logger.debug(LOG_MESSAGES['function_completed'].format(func="generate_reward_script", params="args=(), kwargs=[]", time="0.0000"))
    
    # Возвращаем информацию о сгенерированных скриптах
    return generated_scripts

def generate_profile_script(data_list=None):
    """
    Генерация скрипта для получения профилей сотрудников
    
    Args:
        data_list (list, optional): Список ID профилей
        
    Returns:
        str: Сгенерированный JavaScript скрипт
    """
    logger = get_script_logger("profile", "generation")
    logger.warning("⚠️ Функция generate_profile_script не реализована")
    return "// Заглушка: функция generate_profile_script не реализована"

def generate_news_details_script(data_list=None):
    """
    Генерация скрипта для получения детальной карточки новости
    
    Args:
        data_list (list, optional): Список ID новостей
        
    Returns:
        str: Сгенерированный JavaScript скрипт
    """
    logger = get_script_logger("news_details", "generation")
    logger.warning("⚠️ Функция generate_news_details_script не реализована")
    return "// Заглушка: функция generate_news_details_script не реализована"

def generate_address_book_tn_script(data_list=None):
    """
    Генерация скрипта для получения карточки сотрудника по табельному номеру
    
    Args:
        data_list (list, optional): Список табельных номеров
        
    Returns:
        str: Сгенерированный JavaScript скрипт
    """
    logger = get_script_logger("address_book_tn", "generation")
    logger.warning("⚠️ Функция generate_address_book_tn_script не реализована")
    return "// Заглушка: функция generate_address_book_tn_script не реализована"

def generate_address_book_dev_script(data_list=None):
    """
    Генерация скрипта для получения карточки подразделения
    
    Args:
        data_list (list, optional): Список ID подразделений
        
    Returns:
        str: Сгенерированный JavaScript скрипт
    """
    logger = get_script_logger("address_book_dev", "generation")
    logger.warning("⚠️ Функция generate_address_book_dev_script не реализована")
    return "// Заглушка: функция generate_address_book_dev_script не реализована"

def generate_orders_script(data_list=None):
    """
    Генерация скрипта для получения списка сотрудников с преференциями
    
    Args:
        data_list (list, optional): Список ID сотрудников
        
    Returns:
        str: Сгенерированный JavaScript скрипт
    """
    logger = get_script_logger("orders", "generation")
    logger.warning("⚠️ Функция generate_orders_script не реализована")
    return "// Заглушка: функция generate_orders_script не реализована"

def generate_news_list_script(data_list=None):
    """
    Генерация скрипта для получения списка новостей
    
    Args:
        data_list (list, optional): Список категорий новостей
        
    Returns:
        str: Сгенерированный JavaScript скрипт
    """
    logger = get_script_logger("news_list", "generation")
    logger.warning("⚠️ Функция generate_news_list_script не реализована")
    return "// Заглушка: функция generate_news_list_script не реализована"

def generate_rating_list_script(data_list=None):
    """
    Генерация скрипта для получения рейтинга участников
    
    Args:
        data_list (list, optional): Список ID участников
        
    Returns:
        str: Сгенерированный JavaScript скрипт
    """
    logger = get_script_logger("rating_list", "generation")
    logger.warning("⚠️ Функция generate_rating_list_script не реализована")
    return "// Заглушка: функция generate_rating_list_script не реализована"

# =============================================================================
# ФУНКЦИИ ОБРАБОТКИ JSON В EXCEL
# =============================================================================

def load_json_data(input_json_path):
    """
    Общая функция для загрузки JSON данных
    
    Args:
        input_json_path (str): Путь к входному JSON файлу
        
    Returns:
        dict: Загруженные JSON данные
    """
    try:
        logger.info(LOG_MESSAGES['json_data_loading'])
        logger.debug(f"Загружаем JSON файл: {input_json_path}")
        with open(input_json_path, 'r', encoding='utf-8') as f:
            json_data = json.load(f)
        logger.debug(f"JSON загружен. Тип: {type(json_data)}, количество ключей: {len(json_data) if isinstance(json_data, dict) else 'не dict'}")
        return json_data
    except Exception as e:
        logger.error(LOG_MESSAGES['json_load_error'].format(file_path=input_json_path, error=e))
        return None

@measure_time
def convert_to_integer(value, column_name=None):
    """
    Преобразование значения в целое число с очисткой от лишних символов
    
    Args:
        value: Значение для преобразования
        column_name (str, optional): Имя колонки для логирования
        
    Returns:
        int: Преобразованное целое число или 0 при ошибке
    """
    import re
    
    try:
        if pd.isna(value) or value == '':
            return 0
        
        # Преобразуем в строку и очищаем от лишних символов
        value_str = str(value).strip()
        
        # Удаляем все символы кроме цифр, точки, запятой и минуса
        cleaned_value = re.sub(r'[^\d.,\-]', '', value_str)
        
        # Заменяем запятую на точку для корректного парсинга
        cleaned_value = cleaned_value.replace(',', '.')
        
        # Удаляем лишние точки (оставляем только первую)
        if cleaned_value.count('.') > 1:
            parts = cleaned_value.split('.')
            cleaned_value = parts[0] + '.' + ''.join(parts[1:])
        
        # Преобразуем в число и округляем до целого
        numeric_value = float(cleaned_value)
        return int(numeric_value)
        
    except Exception as e:
        context = f" для колонки {column_name}" if column_name else ""
        logger.warning(LOG_MESSAGES['numeric_conversion_error'].format(
            column=column_name or 'unknown', 
            value=str(value), 
            error=str(e)
        ))
        return 0

@measure_time
def convert_to_float(value, decimal_places=2, column_name=None):
    """
    Преобразование значения в дробное число с очисткой от лишних символов
    
    Args:
        value: Значение для преобразования
        decimal_places (int): Количество знаков после запятой
        column_name (str, optional): Имя колонки для логирования
        
    Returns:
        float: Преобразованное дробное число или 0.0 при ошибке
    """
    import re
    
    try:
        if pd.isna(value) or value == '':
            return 0.0
        
        # Преобразуем в строку и очищаем от лишних символов
        value_str = str(value).strip()
        
        # Удаляем все символы кроме цифр, точки, запятой и минуса
        cleaned_value = re.sub(r'[^\d.,\-]', '', value_str)
        
        # Заменяем запятую на точку для корректного парсинга
        cleaned_value = cleaned_value.replace(',', '.')
        
        # Удаляем лишние точки (оставляем только первую)
        if cleaned_value.count('.') > 1:
            parts = cleaned_value.split('.')
            cleaned_value = parts[0] + '.' + ''.join(parts[1:])
        
        # Преобразуем в число и округляем
        numeric_value = float(cleaned_value)
        return round(numeric_value, decimal_places)
        
    except Exception as e:
        context = f" для колонки {column_name}" if column_name else ""
        logger.warning(LOG_MESSAGES['numeric_conversion_error'].format(
            column=column_name or 'unknown', 
            value=str(value), 
            error=str(e)
        ))
        return 0.0

@measure_time
def convert_to_date(value, input_format='DD.MM.YY', column_name=None):
    """
    Преобразование значения в дату с очисткой от лишних символов
    
    Args:
        value: Значение для преобразования
        input_format (str): Входной формат даты
        column_name (str, optional): Имя колонки для логирования
        
    Returns:
        datetime: Преобразованная дата или исходное значение при ошибке
    """
    from datetime import datetime
    import re
    
    try:
        if pd.isna(value) or value == '':
            return ''
        
        # Преобразуем в строку и очищаем от лишних символов
        value_str = str(value).strip()
        
        # Удаляем лишние пробелы, табы и другие символы
        cleaned_value = re.sub(r'\s+', '', value_str)
        
        # Конвертируем формат в Python datetime format
        python_input_format = input_format.replace('DD', '%d').replace('MM', '%m').replace('YY', '%y').replace('YYYY', '%Y')
        
        # Парсим дату
        parsed_date = datetime.strptime(cleaned_value, python_input_format)
        return parsed_date
        
    except Exception as e:
        logger.warning(LOG_MESSAGES['date_conversion_error'].format(
            column=column_name or 'unknown', 
            value=str(value), 
            error=str(e)
        ))
        return str(value)  # Возвращаем исходное значение если не удалось преобразовать

def apply_column_settings(df, column_settings):
    """
    Применение настроек колонок к DataFrame
    
    Args:
        df (pd.DataFrame): DataFrame для обработки
        column_settings (dict): Настройки колонок
        
    Returns:
        pd.DataFrame: Обработанный DataFrame
    """
    import pandas as pd
    
    df_result = df.copy()
    
    # 1. Применяем преобразования типов данных перед фильтрацией колонок
    
    # Преобразования в числовой формат
    numeric_conversions = column_settings.get('numeric_conversions', {})
    for group_name, group_settings in numeric_conversions.items():
        fields = group_settings.get('fields', [])
        conversion_type = group_settings.get('type', 'integer')
        replace_original = group_settings.get('replace_original', True)
        decimal_places = group_settings.get('decimal_places', 2)
        
        logger.info(LOG_MESSAGES['group_conversion_start'].format(
            group=group_name, 
            type=conversion_type,
            fields=fields
        ))
        
        for column in fields:
            if column in df_result.columns:
                try:
                    logger.info(LOG_MESSAGES['column_conversion_start'].format(
                        column=column, 
                        type=conversion_type
                    ))
                    
                    if conversion_type == 'integer':
                        # Преобразуем в целое число
                        new_values = df_result[column].apply(lambda x: convert_to_integer(x, column))
                    elif conversion_type == 'float':
                        # Преобразуем в дробное число
                        new_values = df_result[column].apply(lambda x: convert_to_float(x, decimal_places, column))
                    
                    if replace_original:
                        df_result[column] = new_values
                        logger.info(LOG_MESSAGES['column_conversion_success'].format(
                            column=column, 
                            type=conversion_type, 
                            action="заменено"
                        ))
                    else:
                        new_column_name = f"{column}_numeric"
                        df_result[new_column_name] = new_values
                        logger.info(LOG_MESSAGES['column_conversion_success'].format(
                            column=column, 
                            type=conversion_type, 
                            action=f"создано новое поле {new_column_name}"
                        ))
                        
                except Exception as e:
                    logger.error(LOG_MESSAGES['column_conversion_failed'].format(
                        column=column, 
                        type=conversion_type, 
                        error=str(e)
                    ))
            else:
                logger.warning(LOG_MESSAGES['column_not_found'].format(
                    column=column,
                    group=group_name
                ))
        
        logger.info(LOG_MESSAGES['group_conversion_completed'].format(
            group=group_name,
            processed=len([col for col in fields if col in df_result.columns])
        ))
    
    # Преобразования дат
    date_conversions = column_settings.get('date_conversions', {})
    for column, settings in date_conversions.items():
        if column in df_result.columns:
            input_format = settings.get('input_format', 'DD.MM.YY')
            replace_original = settings.get('replace_original', True)
            
            try:
                logger.info(LOG_MESSAGES['column_conversion_start'].format(
                    column=column, 
                    type="date"
                ))
                
                new_values = df_result[column].apply(lambda x: convert_to_date(x, input_format, column))
                
                if replace_original:
                    df_result[column] = new_values
                    logger.info(LOG_MESSAGES['column_conversion_success'].format(
                        column=column, 
                        type="date", 
                        action="заменено"
                    ))
                else:
                    new_column_name = f"{column}_formatted"
                    df_result[new_column_name] = new_values
                    logger.info(LOG_MESSAGES['column_conversion_success'].format(
                        column=column, 
                        type="date", 
                        action=f"создано новое поле {new_column_name}"
                    ))
                    
            except Exception as e:
                logger.error(LOG_MESSAGES['column_conversion_failed'].format(
                    column=column, 
                    type="date", 
                    error=str(e)
                ))
    
    # 2. Фильтрация колонок
    columns_to_keep = column_settings.get('columns_to_keep', [])
    columns_to_remove = column_settings.get('columns_to_remove', [])
    
    # Если указаны колонки для сохранения, оставляем только их
    if columns_to_keep:
        # Оставляем только указанные колонки (если они существуют)
        existing_columns_to_keep = [col for col in columns_to_keep if col in df_result.columns]
        df_result = df_result[existing_columns_to_keep]
        logger.info(LOG_MESSAGES['columns_filtered'].format(
            action="оставлены", 
            count=len(existing_columns_to_keep), 
            columns=existing_columns_to_keep
        ))
    
    # Удаляем ненужные колонки
    if columns_to_remove:
        columns_to_drop = [col for col in columns_to_remove if col in df_result.columns]
        df_result = df_result.drop(columns=columns_to_drop)
        logger.info(LOG_MESSAGES['columns_filtered'].format(
            action="удалены", 
            count=len(columns_to_drop), 
            columns=columns_to_drop
        ))
    
    return df_result

def apply_cell_formatting(workbook, df, config_key=None):
    """
    Применение форматирования ячеек в Excel
    
    Args:
        workbook: Объект рабочей книги Excel
        df (DataFrame): DataFrame с данными
        config_key (str, optional): Ключ конфигурации для получения настроек
    """
    from openpyxl.styles import NamedStyle
    from openpyxl.utils import get_column_letter
    
    try:
        # Получаем настройки форматирования из конфигурации
        column_settings = {}
        if config_key and config_key in FUNCTION_CONFIGS:
            config = FUNCTION_CONFIGS[config_key]
            if config_key == "reward" and "reward_profiles" in config:
                column_settings = config["reward_profiles"].get("column_settings", {})
            elif config_key == "leaders_for_admin" and "leaders_processing" in config:
                column_settings = config["leaders_processing"].get("column_settings", {})
        
        if not column_settings:
            return
        
        # Получаем лист DATA
        worksheet = workbook['DATA']
        
        # Создаем стили для форматирования
        number_style = NamedStyle(name="number_style")
        number_style.number_format = '#,##0'
        
        float_style = NamedStyle(name="float_style")
        float_style.number_format = '#,##0.00'
        
        date_style = NamedStyle(name="date_style")
        date_style.number_format = 'YYYY-MM-DD'
        
        # Применяем форматирование к числовым колонкам
        numeric_conversions = column_settings.get('numeric_conversions', {})
        for group_name, group_settings in numeric_conversions.items():
            fields = group_settings.get('fields', [])
            conversion_type = group_settings.get('type', 'integer')
            decimal_places = group_settings.get('decimal_places', 2)
            
            for column in fields:
                if column in df.columns:
                    col_idx = df.columns.get_loc(column) + 1  # +1 потому что Excel начинается с 1
                    col_letter = get_column_letter(col_idx)
                    
                    if conversion_type == 'integer':
                        # Применяем целочисленный формат
                        for row in range(2, len(df) + 2):  # +2 потому что Excel начинается с 1 и есть заголовок
                            cell = worksheet[f'{col_letter}{row}']
                            cell.style = number_style
                    elif conversion_type == 'float':
                        # Применяем дробный формат
                        float_style.number_format = f'#,##0.{"0" * decimal_places}'
                        for row in range(2, len(df) + 2):
                            cell = worksheet[f'{col_letter}{row}']
                            cell.style = float_style
        
        # Применяем форматирование к датам
        date_conversions = column_settings.get('date_conversions', {})
        for column, settings in date_conversions.items():
            if column in df.columns:
                col_idx = df.columns.get_loc(column) + 1
                col_letter = get_column_letter(col_idx)
                
                # Применяем формат даты
                for row in range(2, len(df) + 2):
                    cell = worksheet[f'{col_letter}{row}']
                    cell.style = date_style
        
        # Применяем форматирование к новым колонкам (с суффиксами)
        for column in df.columns:
            if column.endswith('_numeric'):
                col_idx = df.columns.get_loc(column) + 1
                col_letter = get_column_letter(col_idx)
                
                # Определяем тип по исходной колонке
                original_column = column.replace('_numeric', '')
                
                # Ищем исходную колонку в группах
                found_conversion_type = None
                found_decimal_places = 2
                
                for group_name, group_settings in numeric_conversions.items():
                    fields = group_settings.get('fields', [])
                    if original_column in fields:
                        found_conversion_type = group_settings.get('type', 'integer')
                        found_decimal_places = group_settings.get('decimal_places', 2)
                        break
                
                if found_conversion_type:
                    if found_conversion_type == 'integer':
                        for row in range(2, len(df) + 2):
                            cell = worksheet[f'{col_letter}{row}']
                            cell.style = number_style
                    elif found_conversion_type == 'float':
                        float_style.number_format = f'#,##0.{"0" * found_decimal_places}'
                        for row in range(2, len(df) + 2):
                            cell = worksheet[f'{col_letter}{row}']
                            cell.style = float_style
            
            elif column.endswith('_formatted'):
                col_idx = df.columns.get_loc(column) + 1
                col_letter = get_column_letter(col_idx)
                
                # Применяем формат даты к отформатированным колонкам
                for row in range(2, len(df) + 2):
                    cell = worksheet[f'{col_letter}{row}']
                    cell.style = date_style
                    
    except Exception as e:
        logger.warning(f"Ошибка при применении форматирования ячеек: {e}")

def save_excel_file(df, output_excel_path, config_key=None):
    """
    Общая функция для сохранения DataFrame в Excel с применением стилей
    
    Args:
        df (DataFrame): DataFrame для сохранения
        output_excel_path (str): Путь к выходному Excel файлу
        config_key (str, optional): Ключ конфигурации для получения настроек
        
    Returns:
        bool: True если сохранение успешно, False в противном случае
    """
    try:
        # Создание директории для выходного файла если не существует
        output_dir = os.path.dirname(output_excel_path)
        if output_dir and not os.path.exists(output_dir):
            os.makedirs(output_dir)
            logger.info(LOG_MESSAGES['json_directory_created'].format(directory=output_dir))
        
        # Создание Excel файла
        logger.info(LOG_MESSAGES['json_excel_creation'])
        with pd.ExcelWriter(output_excel_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='DATA', index=False)
            workbook = writer.book
            
            # Получаем настройки закрепления из конфигурации
            freeze_cell = "B2"  # По умолчанию закрепляем первую строку и первую колонку
            if config_key and config_key in FUNCTION_CONFIGS:
                config = FUNCTION_CONFIGS[config_key]
                # Проверяем, есть ли вложенные конфигурации
                if config_key == "reward" and "reward_profiles" in config:
                    reward_profiles_config = config["reward_profiles"]
                    freeze_cell = reward_profiles_config.get('excel_freeze_cell', "B2")
                elif config_key == "leaders_for_admin" and "leaders_processing" in config:
                    leaders_processing_config = config["leaders_processing"]
                    freeze_cell = leaders_processing_config.get('excel_freeze_cell', "B2")
                else:
                    freeze_cell = config.get('excel_freeze_cell', "B2")
            
            # Применяем стили с настройками закрепления
            apply_excel_styling(workbook, freeze_cell)
            
            # Применяем форматирование ячеек
            apply_cell_formatting(workbook, df, config_key)
            
            # Создание дополнительных листов
            create_summary_sheet(workbook, df)
            create_statistics_sheet(workbook, df)
            
            # Создание специального листа для reward данных
            if config_key == "reward" or (config_key and "reward" in config_key):
                create_reward_summary_sheet(workbook, df)
        
        logger.info(LOG_MESSAGES['json_excel_success'].format(file_path=output_excel_path))
        return True
        
    except Exception as e:
        logger.error(LOG_MESSAGES['excel_creation_error'].format(error=e))
        return False

@measure_time
def convert_leaders_json_to_excel(input_json_path, output_excel_path, config_key=None):
    """
    Конвертация JSON файла с данными лидеров в Excel
    
    Args:
        input_json_path (str): Путь к входному JSON файлу
        output_excel_path (str): Путь к выходному Excel файлу
        config_key (str, optional): Ключ конфигурации для получения настроек
        
    Returns:
        bool: True если конвертация успешна, False в противном случае
    """
    try:
        script_logger = get_script_logger("leaders_for_admin", "conversion")
        script_logger.info(LOG_MESSAGES['json_conversion_start'].format(input=input_json_path, output=output_excel_path))
        
        # Проверка существования входного файла
        if not os.path.exists(input_json_path):
            logger.error(LOG_MESSAGES['json_file_not_found'].format(file_path=input_json_path))
            return False
        
        # Загрузка JSON данных
        json_data = load_json_data(input_json_path)
        if json_data is None:
            return False
        
        # Обработка данных
        logger.info(LOG_MESSAGES['json_data_processing'])
        all_leaders_data = []
        
        if isinstance(json_data, dict):
            # Обрабатываем все турниры в структуре LeadersForAdmin
            total_tournaments = 0
            total_leaders = 0
            
            for tournament_key, tournament_value in json_data.items():
                if isinstance(tournament_value, list) and len(tournament_value) > 0:
                    # Проверяем, содержит ли первый элемент данные о турнире
                    first_item = tournament_value[0]
                    if isinstance(first_item, dict) and 'body' in first_item:
                        body = first_item['body']
                        if 'tournament' in body:
                            tournament = body['tournament']
                            if 'leaders' in tournament:
                                tournament_leaders = tournament['leaders']
                                if tournament_leaders:
                                    # Добавляем информацию о турнире к каждому лидеру
                                    for leader in tournament_leaders:
                                        leader_with_tournament = leader.copy()
                                        leader_with_tournament['tournamentId'] = tournament.get('tournamentId', tournament_key)
                                        leader_with_tournament['tournamentIndicator'] = tournament.get('tournamentIndicator', '')
                                        leader_with_tournament['tournamentStatus'] = tournament.get('status', '')
                                        leader_with_tournament['contestants'] = tournament.get('contestants', '')
                                        all_leaders_data.append(leader_with_tournament)
                                    
                                    total_tournaments += 1
                                    total_leaders += len(tournament_leaders)
                                    logger.debug(LOG_MESSAGES['json_leaders_found'].format(key=tournament_key, count=len(tournament_leaders)))
            
            logger.info(LOG_MESSAGES['tournaments_processed'].format(tournaments=total_tournaments, leaders=total_leaders))
            leaders_data = all_leaders_data
            
        elif isinstance(json_data, list):
            # Прямой список лидеров
            leaders_data = json_data
            logger.info(LOG_MESSAGES['json_direct_leaders'].format(count=len(leaders_data)))
        else:
            logger.error(LOG_MESSAGES['json_invalid_format'])
            return False
        
        if not leaders_data:
            logger.error(LOG_MESSAGES['json_no_leaders'])
            return False
        
        # Преобразование данных в плоскую структуру
        flattened_data = []
        for leader in leaders_data:
            flattened_leader = flatten_leader_data(leader)
            flattened_data.append(flattened_leader)
        
        # Создание DataFrame
        df = pd.DataFrame(flattened_data)
        
        if df.empty:
            logger.warning(LOG_MESSAGES['no_data_warning'])
            return False
        
        logger.info(LOG_MESSAGES['json_records_processed'].format(count=len(df)))
        
        # Применение настроек колонок
        if config_key == "leaders_for_admin" and "leaders_for_admin" in FUNCTION_CONFIGS:
            config = FUNCTION_CONFIGS["leaders_for_admin"]
            if "leaders_processing" in config and "column_settings" in config["leaders_processing"]:
                column_settings = config["leaders_processing"]["column_settings"]
                logger.info(LOG_MESSAGES['column_settings_applying_leaders'])
                df = apply_column_settings(df, column_settings)
                logger.info(LOG_MESSAGES['column_settings_applied'].format(count=len(df.columns)))
        
        # Сохранение в Excel
        return save_excel_file(df, output_excel_path, "leaders_for_admin")
        
    except Exception as e:
        logger.error(LOG_MESSAGES['json_leaders_conversion_error'].format(error=e))
        return False

def extract_profiles_from_data(data, structure):
    """
    Извлечение профилей из данных API наград
    
    Args:
        data (dict): Данные ответа API
        structure (str): Тип структуры данных
        
    Returns:
        list: Список профилей или None
    """
    try:
        if structure == 'body.badge.profiles':
            # Структура 1: body.badge.profiles
            if data.get('body', {}).get('badge', {}).get('profiles'):
                return data['body']['badge']['profiles']
        elif structure == 'body.profiles':
            # Структура 2: body.profiles
            if data.get('body', {}).get('profiles'):
                return data['body']['profiles']
        elif structure == 'body.array':
            # Структура 3: прямой массив профилей в body
            if isinstance(data.get('body'), list):
                return data['body']
        elif structure == 'root.array':
            # Структура 4: прямой массив в корне
            if isinstance(data, list):
                return data
        
        # Попытка автоматического определения структуры
        if data.get('body', {}).get('badge', {}).get('profiles'):
            return data['body']['badge']['profiles']
        elif data.get('body', {}).get('profiles'):
            return data['body']['profiles']
        elif isinstance(data.get('body'), list):
            return data['body']
        elif isinstance(data, list):
            return data
        
        return None
    except Exception as e:
        logger.error(LOG_MESSAGES['profile_extraction_error'].format(error=e))
        return None

@measure_time
def convert_reward_profiles_json_to_excel(input_json_path, output_excel_path, config_key=None):
    """
    Конвертация JSON файла с данными профилей наград в Excel
    
    Args:
        input_json_path (str): Путь к входному JSON файлу
        output_excel_path (str): Путь к выходному Excel файлу
        config_key (str, optional): Ключ конфигурации для получения настроек
        
    Returns:
        bool: True если конвертация успешна, False в противном случае
    """
    try:
        logger.info(LOG_MESSAGES['json_conversion_start'].format(input=input_json_path, output=output_excel_path))
        
        # Проверка существования входного файла
        if not os.path.exists(input_json_path):
            logger.error(LOG_MESSAGES['json_file_not_found'].format(file_path=input_json_path))
            return False
        
        # Загрузка JSON данных
        json_data = load_json_data(input_json_path)
        if json_data is None:
            return False
        
        # Обработка данных
        logger.info(LOG_MESSAGES['json_data_processing'])
        all_leaders_data = []
        
        if isinstance(json_data, dict):
            # Обрабатываем все коды наград
            total_rewards = 0
            total_leaders = 0
            
            for reward_code, reward_value in json_data.items():
                # Новая структура: массив с body.badge.leaders
                if isinstance(reward_value, list) and reward_value:
                    first_item = reward_value[0]
                    if isinstance(first_item, dict) and 'body' in first_item:
                        body = first_item['body']
                        if isinstance(body, dict) and 'badge' in body:
                            badge = body['badge']
                            
                            # Получаем лидеров из body.badge.leaders
                            leaders = badge.get('leaders', [])
                            contestants = badge.get('contestants', '')
                            badge_id = badge.get('badgeId', reward_code)
                            
                            if leaders:
                                # Добавляем информацию о коде награды к каждому лидеру
                                for leader in leaders:
                                    if isinstance(leader, dict):
                                        leader_with_reward = flatten_reward_leader_data(leader, reward_code)
                                        
                                        # Добавляем информацию о награде
                                        leader_with_reward['badgeId'] = badge_id
                                        leader_with_reward['contestants'] = contestants
                                        leader_with_reward['profilesCount'] = len(leaders)
                                        
                                        all_leaders_data.append(leader_with_reward)
                                
                                total_rewards += 1
                                total_leaders += len(leaders)
                                logger.debug(LOG_MESSAGES['json_reward_found'].format(key=reward_code, count=len(leaders)))
                                logger.info(LOG_MESSAGES['reward_profiles_leaders_found'].format(code=reward_code, count=len(leaders), structure="body.badge.leaders"))
                
                # Старая структура: объект с badgeInfo.leaders
                elif isinstance(reward_value, dict):
                    # Получаем информацию о награде
                    profiles_count = reward_value.get('profilesCount', 0)
                    badge_info = reward_value.get('badgeInfo', {})
                    contestants = badge_info.get('contestants', '')
                    
                    # Получаем лидеров из badgeInfo.leaders
                    leaders = badge_info.get('leaders', [])
                    
                    if leaders:
                        # Добавляем информацию о коде награды к каждому лидеру
                        for leader in leaders:
                            if isinstance(leader, dict):
                                leader_with_reward = flatten_reward_leader_data(leader, reward_code)
                                
                                # Добавляем информацию о награде
                                leader_with_reward['badgeId'] = badge_info.get('badgeId', '')
                                leader_with_reward['contestants'] = contestants
                                leader_with_reward['profilesCount'] = profiles_count
                                
                                all_leaders_data.append(leader_with_reward)
                        
                        total_rewards += 1
                        total_leaders += len(leaders)
                        logger.debug(LOG_MESSAGES['json_reward_found'].format(key=reward_code, count=len(leaders)))
                        logger.info(LOG_MESSAGES['reward_profiles_leaders_found'].format(code=reward_code, count=len(leaders), structure="badgeInfo.leaders"))
            
            logger.info(LOG_MESSAGES['reward_profiles_leaders_processed'].format(rewards=total_rewards, leaders=total_leaders))
            leaders_data = all_leaders_data
            
        elif isinstance(json_data, list):
            # Прямой список лидеров
            leaders_data = json_data
            logger.info(LOG_MESSAGES['json_direct_leaders'].format(count=len(leaders_data)))
        else:
            logger.error(LOG_MESSAGES['json_invalid_format'])
            return False
        
        if not leaders_data:
            logger.error(LOG_MESSAGES['no_profiles_error'])
            return False
        
        # Создание DataFrame
        df = pd.DataFrame(leaders_data)
        
        if df.empty:
            logger.warning(LOG_MESSAGES['no_data_warning'])
            return False
        
        logger.info(LOG_MESSAGES['json_records_processed'].format(count=len(df)))
        
        # Применение настроек колонок
        if config_key == "reward" and "reward" in FUNCTION_CONFIGS:
            config = FUNCTION_CONFIGS["reward"]
            if "reward_profiles" in config and "column_settings" in config["reward_profiles"]:
                column_settings = config["reward_profiles"]["column_settings"]
                logger.info(LOG_MESSAGES['column_settings_applying'])
                df = apply_column_settings(df, column_settings)
                logger.info(LOG_MESSAGES['column_settings_applied'].format(count=len(df.columns)))
        
        # Сохранение в Excel
        return save_excel_file(df, output_excel_path, "reward")
        
    except Exception as e:
        logger.error(LOG_MESSAGES['json_reward_profiles_conversion_error'].format(error=e))
        return False

@measure_time
def convert_reward_json_to_excel(input_json_path, output_excel_path, config_key=None):
    """
    Конвертация JSON файла с данными наград в Excel
    
    Args:
        input_json_path (str): Путь к входному JSON файлу
        output_excel_path (str): Путь к выходному Excel файлу
        config_key (str, optional): Ключ конфигурации для получения настроек
        
    Returns:
        bool: True если конвертация успешна, False в противном случае
    """
    try:
        script_logger = get_script_logger("reward", "conversion")
        script_logger.info(LOG_MESSAGES['json_conversion_start'].format(input=input_json_path, output=output_excel_path))
        
        # Проверка существования входного файла
        if not os.path.exists(input_json_path):
            logger.error(LOG_MESSAGES['json_file_not_found'].format(file_path=input_json_path))
            return False
        
        # Загрузка JSON данных
        json_data = load_json_data(input_json_path)
        if json_data is None:
            return False
        
        # Обработка данных
        script_logger.info(LOG_MESSAGES['json_data_processing'])
        script_logger.debug(f"Тип данных: {type(json_data)}")
        script_logger.debug(f"Ключи в данных: {list(json_data.keys()) if isinstance(json_data, dict) else 'не словарь'}")
        all_profiles_data = []
        
        if isinstance(json_data, dict):
            # Обрабатываем все коды наград
            total_rewards = 0
            total_profiles = 0
            
            script_logger.debug(f"Начинаем обработку {len(json_data)} кодов наград")
            
            for reward_code, reward_value in json_data.items():
                script_logger.debug(f"Обрабатываем код награды: {reward_code}, тип значения: {type(reward_value)}")
                # Новая структура данных (список как в leaders)
                if isinstance(reward_value, list) and len(reward_value) > 0:
                    first_data = reward_value[0]
                    if isinstance(first_data, dict) and 'body' in first_data:
                        body = first_data.get('body', {})
                        badge = body.get('badge', {})
                        profiles = badge.get('profiles', [])
                        badge_info = badge
                        
                        script_logger.debug(f"Обрабатываем структуру массива для {reward_code}: профилей={len(profiles)}")
                        
                        if profiles and len(profiles) > 0:
                            # Добавляем информацию о коде награды и данных награды к каждому профилю
                            for profile in profiles:
                                if isinstance(profile, dict):
                                    profile_with_reward = profile.copy()
                                    profile_with_reward['rewardCode'] = reward_code
                                    
                                    # Добавляем информацию о награде
                                    if badge_info:
                                        profile_with_reward['badgeName'] = badge_info.get('name', '')
                                        profile_with_reward['badgeDescription'] = badge_info.get('description', '')
                                        profile_with_reward['badgeType'] = badge_info.get('type', '')
                                        profile_with_reward['badgeCategory'] = badge_info.get('category', '')
                                    
                                    all_profiles_data.append(profile_with_reward)
                            
                            total_rewards += 1
                            total_profiles += len(profiles)
                            script_logger.debug(LOG_MESSAGES['json_reward_found'].format(key=reward_code, count=len(profiles)))
                            script_logger.info(f"Найдено профилей для кода награды {reward_code}: {len(profiles)} (структура массива)")
                        else:
                            script_logger.debug(f"Профили пусты для {reward_code}: {len(profiles)} профилей")
                
                # Старая структура данных (прямая структура с profiles)
                elif isinstance(reward_value, dict) and 'profiles' in reward_value:
                    profiles = reward_value.get('profiles', [])
                    profiles_count = reward_value.get('profilesCount', 0)
                    badge_info = reward_value.get('badgeInfo', {})
                    
                    script_logger.debug(f"Обрабатываем прямую структуру для {reward_code}: профилей={len(profiles)}")
                    
                    if profiles and len(profiles) > 0:
                        # Добавляем информацию о коде награды и данных награды к каждому профилю
                        for profile in profiles:
                            if isinstance(profile, dict):
                                profile_with_reward = profile.copy()
                                profile_with_reward['rewardCode'] = reward_code
                                
                                # Добавляем информацию о награде
                                if badge_info:
                                    profile_with_reward['badgeName'] = badge_info.get('name', '')
                                    profile_with_reward['badgeDescription'] = badge_info.get('description', '')
                                    profile_with_reward['badgeType'] = badge_info.get('type', '')
                                    profile_with_reward['badgeCategory'] = badge_info.get('category', '')
                                
                                all_profiles_data.append(profile_with_reward)
                        
                        total_rewards += 1
                        total_profiles += len(profiles)
                        script_logger.debug(LOG_MESSAGES['json_reward_found'].format(key=reward_code, count=len(profiles)))
                        script_logger.info(f"Найдено профилей для кода награды {reward_code}: {len(profiles)} (прямая структура)")
                    else:
                        script_logger.debug(f"Профили пусты для {reward_code}: {len(profiles)} профилей")
                
                # Новая структура данных (с информацией о структуре - старая логика)
                elif isinstance(reward_value, dict) and ('data' in reward_value or 'structure' in reward_value):
                    data = reward_value.get('data', {})
                    structure = reward_value.get('structure', 'unknown')
                    profiles_count = reward_value.get('profilesCount', 0)
                    badge_info = reward_value.get('badgeInfo', {})
                    
                    # Извлекаем профили из данных
                    profiles = extract_profiles_from_data(data, structure)
                    
                    if profiles:
                        # Добавляем информацию о коде награды и данных награды к каждому профилю
                        for profile in profiles:
                            if isinstance(profile, dict):
                                profile_with_reward = profile.copy()
                                profile_with_reward['rewardCode'] = reward_code
                                profile_with_reward['structure'] = structure
                                
                                # Добавляем информацию о награде
                                if badge_info:
                                    profile_with_reward['badgeName'] = badge_info.get('name', '')
                                    profile_with_reward['badgeDescription'] = badge_info.get('description', '')
                                    profile_with_reward['badgeType'] = badge_info.get('type', '')
                                    profile_with_reward['badgeCategory'] = badge_info.get('category', '')
                                
                                all_profiles_data.append(profile_with_reward)
                        
                        total_rewards += 1
                        total_profiles += len(profiles)
                        script_logger.debug(LOG_MESSAGES['json_reward_found'].format(key=reward_code, count=len(profiles)))
                        script_logger.info(LOG_MESSAGES['reward_profiles_found'].format(code=reward_code, count=len(profiles), structure=structure))
                
                # Старая структура данных (для обратной совместимости)
                elif isinstance(reward_value, list) and len(reward_value) > 0:
                    # Проверяем, содержит ли первый элемент данные о награде
                    first_item = reward_value[0]
                    script_logger.debug(f"Обрабатываем reward_value для {reward_code}: тип={type(reward_value)}, длина={len(reward_value)}")
                    script_logger.debug(f"Первый элемент: тип={type(first_item)}, ключи={list(first_item.keys()) if isinstance(first_item, dict) else 'не словарь'}")
                    
                    if isinstance(first_item, dict) and 'body' in first_item:
                        body = first_item['body']
                        script_logger.debug(f"Найден body: ключи={list(body.keys()) if isinstance(body, dict) else 'не словарь'}")
                        
                        # Проверяем разные возможные структуры данных
                        profiles = None
                        badge_info = None
                        
                        # Структура 1: body.badge.profiles
                        if 'badge' in body and 'profiles' in body['badge']:
                            profiles = body['badge']['profiles']
                            badge_info = body['badge']
                            script_logger.debug(f"Найдена структура body.badge.profiles: количество профилей={len(profiles) if profiles else 0}")
                        # Структура 2: body.profiles (прямые профили)
                        elif 'profiles' in body:
                            profiles = body['profiles']
                            badge_info = body
                            script_logger.debug(f"Найдена структура body.profiles: количество профилей={len(profiles) if profiles else 0}")
                        else:
                            script_logger.debug(f"Не найдена структура profiles в body")
                        
                        if profiles and isinstance(profiles, list):
                            # Добавляем информацию о коде награды и данных награды к каждому профилю
                            for profile in profiles:
                                if isinstance(profile, dict):
                                    profile_with_reward = profile.copy()
                                    profile_with_reward['rewardCode'] = reward_code
                                    
                                    # Добавляем информацию о награде
                                    if badge_info:
                                        profile_with_reward['badgeName'] = badge_info.get('name', '')
                                        profile_with_reward['badgeDescription'] = badge_info.get('description', '')
                                        profile_with_reward['badgeType'] = badge_info.get('type', '')
                                        profile_with_reward['badgeCategory'] = badge_info.get('category', '')
                                    
                                    all_profiles_data.append(profile_with_reward)
                            
                            total_rewards += 1
                            total_profiles += len(profiles)
                            script_logger.debug(LOG_MESSAGES['json_reward_found'].format(key=reward_code, count=len(profiles)))
                            script_logger.info(LOG_MESSAGES['reward_profiles_found_old'].format(code=reward_code, count=len(profiles)))
                        else:
                            script_logger.debug(f"Профили не найдены или не являются списком для {reward_code}")
                    else:
                        script_logger.debug(f"Первый элемент не содержит body для {reward_code}")
                
                # Обработка словаря старой структуры (когда reward_value - dict, но без ключей новой структуры)
                elif isinstance(reward_value, dict):
                    script_logger.debug(f"Обрабатываем старую структуру dict для {reward_code}")
                    # Возможно, это старая структура в виде словаря
                    # Ищем профили напрямую в словаре или в подструктурах
                    profiles = None
                    badge_info = None
                    
                    # Пытаемся найти профили в разных возможных местах
                    if 'profiles' in reward_value:
                        profiles = reward_value['profiles']
                        badge_info = reward_value
                        script_logger.debug(f"Найдены профили напрямую в {reward_code}: {len(profiles) if profiles else 0}")
                    elif 'badge' in reward_value and 'profiles' in reward_value['badge']:
                        profiles = reward_value['badge']['profiles']
                        badge_info = reward_value['badge']
                        script_logger.debug(f"Найдены профили в badge для {reward_code}: {len(profiles) if profiles else 0}")
                    elif 'body' in reward_value and 'badge' in reward_value['body'] and 'profiles' in reward_value['body']['badge']:
                        profiles = reward_value['body']['badge']['profiles']
                        badge_info = reward_value['body']['badge']
                        script_logger.debug(f"Найдены профили в body.badge для {reward_code}: {len(profiles) if profiles else 0}")
                    
                    if profiles and isinstance(profiles, list):
                        # Добавляем информацию о коде награды к каждому профилю
                        for profile in profiles:
                            if isinstance(profile, dict):
                                profile_with_reward = profile.copy()
                                profile_with_reward['rewardCode'] = reward_code
                                
                                # Добавляем информацию о награде
                                if badge_info:
                                    profile_with_reward['badgeName'] = badge_info.get('name', '')
                                    profile_with_reward['badgeDescription'] = badge_info.get('description', '')
                                    profile_with_reward['badgeType'] = badge_info.get('type', '')
                                    profile_with_reward['badgeCategory'] = badge_info.get('category', '')
                                
                                all_profiles_data.append(profile_with_reward)
                        
                        total_rewards += 1
                        total_profiles += len(profiles)
                        script_logger.debug(LOG_MESSAGES['json_reward_found'].format(key=reward_code, count=len(profiles)))
                        script_logger.info(LOG_MESSAGES['reward_profiles_found_old'].format(code=reward_code, count=len(profiles)))
                    else:
                        script_logger.debug(f"Профили не найдены в структуре dict для {reward_code}")
                
                # Прямая структура данных (профили в корне)
                elif isinstance(reward_value, list):
                    # Прямой список профилей
                    profiles = reward_value
                    if profiles and isinstance(profiles, list):
                        # Добавляем информацию о коде награды к каждому профилю
                        for profile in profiles:
                            if isinstance(profile, dict):
                                profile_with_reward = profile.copy()
                                profile_with_reward['rewardCode'] = reward_code
                                all_profiles_data.append(profile_with_reward)
                        
                        total_rewards += 1
                        total_profiles += len(profiles)
                        script_logger.debug(LOG_MESSAGES['json_reward_found'].format(key=reward_code, count=len(profiles)))
                        script_logger.info(LOG_MESSAGES['reward_profiles_found_old'].format(code=reward_code, count=len(profiles)))
            
            script_logger.info(LOG_MESSAGES['rewards_processed'].format(rewards=total_rewards, profiles=total_profiles))
            profiles_data = all_profiles_data
            
        elif isinstance(json_data, list):
            # Прямой список профилей
            profiles_data = json_data
            script_logger.info(LOG_MESSAGES['direct_profiles_list'].format(count=len(profiles_data)))
        else:
            script_logger.error(LOG_MESSAGES['json_invalid_format'])
            return False
        
        if not profiles_data:
            script_logger.error(LOG_MESSAGES['no_profiles_error'])
            return False
        
        # Преобразование данных в плоскую структуру
        flattened_data = []
        for profile in profiles_data:
            flattened_profile = flatten_reward_profile_data(profile)
            flattened_data.append(flattened_profile)
        
        # Создание DataFrame
        df = pd.DataFrame(flattened_data)
        
        if df.empty:
            logger.warning(LOG_MESSAGES['no_data_warning'])
            return False
        
        script_logger.info(LOG_MESSAGES['json_records_processed'].format(count=len(df)))
        
        # Сохранение в Excel
        return save_excel_file(df, output_excel_path, "reward")
        
    except Exception as e:
        script_logger.error(LOG_MESSAGES['json_reward_conversion_error'].format(error=e))
        return False

@measure_time
def convert_json_to_excel(input_json_path, output_excel_path, config_key=None):
    """
    Универсальная конвертация JSON файла в Excel (для обратной совместимости)
    
    Args:
        input_json_path (str): Путь к входному JSON файлу
        output_excel_path (str): Путь к выходному Excel файлу
        config_key (str, optional): Ключ конфигурации для получения настроек
        
    Returns:
        bool: True если конвертация успешна, False в противном случае
    """
    # Определяем тип данных по config_key
    if config_key == "leaders_for_admin":
        return convert_leaders_json_to_excel(input_json_path, output_excel_path, config_key)
    elif config_key == "reward":
        return convert_reward_json_to_excel(input_json_path, output_excel_path, config_key)
    elif config_key == "reward_profiles" or (config_key == "reward" and "reward_profiles" in FUNCTION_CONFIGS["reward"]):
        return convert_reward_profiles_json_to_excel(input_json_path, output_excel_path, "reward")
    elif config_key == "leaders_processing" or (config_key == "leaders_for_admin" and "leaders_processing" in FUNCTION_CONFIGS["leaders_for_admin"]):
        return convert_leaders_json_to_excel(input_json_path, output_excel_path, "leaders_for_admin")
    else:
        # Автоматическое определение по настройкам конфигурации
        file_name = os.path.basename(input_json_path)
        file_name_without_ext = os.path.splitext(file_name)[0]
        
        # Ищем подходящую конфигурацию по json_file
        for config_key_name, config in FUNCTION_CONFIGS.items():
            if "json_file" in config:
                config_json_file = config["json_file"]
                if config_json_file == file_name_without_ext:
                    logger.info(LOG_MESSAGES['json_file_processing'].format(file_name=file_name) + f" (автоопределение по конфигурации: {config_key_name})")
                    
                    # Определяем тип обработки по config_key_name
                    if config_key_name == "leaders_for_admin":
                        return convert_leaders_json_to_excel(input_json_path, output_excel_path, config_key_name)
                    elif config_key_name == "reward":
                        return convert_reward_json_to_excel(input_json_path, output_excel_path, config_key_name)
                    elif config_key_name == "reward_profiles" or (config_key_name == "reward" and "reward_profiles" in FUNCTION_CONFIGS["reward"]):
                        return convert_reward_profiles_json_to_excel(input_json_path, output_excel_path, "reward")
                    elif config_key_name == "leaders_processing" or (config_key_name == "leaders_for_admin" and "leaders_processing" in FUNCTION_CONFIGS["leaders_for_admin"]):
                        return convert_leaders_json_to_excel(input_json_path, output_excel_path, "leaders_for_admin")
                    else:
                        # Для других типов используем обработку лидеров по умолчанию
                        logger.info(LOG_MESSAGES['json_file_processing'].format(file_name=file_name) + f" (автоопределение: {config_key_name} -> лидеры)")
                        return convert_leaders_json_to_excel(input_json_path, output_excel_path, config_key_name)
        
        # Если конфигурация не найдена, используем автоопределение по имени файла
        file_name_lower = file_name.lower()
        
        # Если файл содержит "profiles" - используем обработку профилей наград
        if "profiles" in file_name_lower:
            logger.info(LOG_MESSAGES['json_file_processing'].format(file_name=file_name) + " (автоопределение по имени: профили наград)")
            return convert_reward_profiles_json_to_excel(input_json_path, output_excel_path, "reward")
        # Если файл содержит "leaders" - используем обработку лидеров
        elif "leaders" in file_name_lower:
            logger.info(LOG_MESSAGES['json_file_processing'].format(file_name=file_name) + " (автоопределение по имени: лидеры)")
            return convert_leaders_json_to_excel(input_json_path, output_excel_path, "leaders_for_admin")
        else:
            # По умолчанию используем обработку лидеров
            logger.info(LOG_MESSAGES['json_file_processing'].format(file_name=file_name) + " (автоопределение по умолчанию: лидеры)")
            return convert_leaders_json_to_excel(input_json_path, output_excel_path, "leaders_for_admin")

@measure_time
def convert_specific_json_file(file_name_without_extension, config_key=None):
    """
    Конвертирует конкретный JSON файл в Excel
    
    Args:
        file_name_without_extension (str): Имя файла без расширения
        config_key (str, optional): Ключ конфигурации для получения настроек Excel
        
    Returns:
        bool: True если конвертация успешна, False в противном случае
    """
    try:
        # Формируем пути к файлам используя новую структуру
        json_dir = os.path.join(BASE_DIR, SUBDIRECTORIES["JSON"])
        output_dir = os.path.join(BASE_DIR, SUBDIRECTORIES["OUTPUT"])
        input_json_path = os.path.join(json_dir, f"{file_name_without_extension}.json")
        
        # Генерируем уникальное имя Excel файла
        if config_key and config_key in FUNCTION_CONFIGS:
            config = FUNCTION_CONFIGS[config_key]
            
            # Проверяем, есть ли вложенные конфигурации
            if config_key == "reward" and "reward_profiles" in config:
                reward_profiles_config = config["reward_profiles"]
                excel_file_base = reward_profiles_config.get("excel_file", file_name_without_extension)
            elif config_key == "leaders_for_admin" and "leaders_processing" in config:
                leaders_processing_config = config["leaders_processing"]
                excel_file_base = leaders_processing_config.get("excel_file", file_name_without_extension)
            else:
                excel_file_base = config.get("excel_file", file_name_without_extension)
            
            # Создаем временную метку
            timestamp = datetime.datetime.now().strftime('%Y-%m-%d-%H-%M-%S')
            
            # Формируем имя файла: имя_из_конфига_YYYY-MM-DD-HH-MM-SS.xlsx (без варианта)
            excel_filename = f"{excel_file_base}_{timestamp}{FILE_EXTENSIONS['EXCEL']}"
        else:
            # Fallback: используем имя JSON файла с временной меткой
            timestamp = datetime.datetime.now().strftime('%Y-%m-%d-%H-%M-%S')
            excel_filename = f"{file_name_without_extension}_{timestamp}{FILE_EXTENSIONS['EXCEL']}"
        
        output_excel_path = os.path.join(output_dir, excel_filename)
        
        logger.info(LOG_MESSAGES['json_file_processing'].format(file_name=file_name_without_extension))
        logger.info(LOG_MESSAGES['excel_file_creation'].format(filename=excel_filename))
        
        # Проверяем существование JSON файла
        if not os.path.exists(input_json_path):
            logger.error(LOG_MESSAGES['json_file_not_found'].format(file_path=input_json_path))
            return False
            
        # Конвертируем файл
        if convert_json_to_excel(input_json_path, output_excel_path, config_key):
            logger.info(LOG_MESSAGES['json_excel_success'].format(file_path=output_excel_path))
            return True
        else:
            return False
            
    except Exception as e:
        logger.error(LOG_MESSAGES['json_conversion_error'].format(error=str(e)))
        return False

# =============================================================================
# ФУНКЦИИ ВЫВОДА СТАТИСТИКИ
# =============================================================================

def print_summary():
    """
    Вывод итоговой статистики работы программы
    
    Формирует и выводит в консоль и лог подробную статистику выполнения:
    - Общее время работы
    - Количество обработанных действий
    - Количество выполненных функций
    - Время выполнения каждой функции
    """
    global program_start_time, processed_actions_count, function_execution_times
    
    total_time = time.time() - program_start_time
    current_time = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]
    
    # Формирование строк статистики
    summary_lines = [
        "=" * 70,
        f"SUMMARY - {LOG_MESSAGES['summary_stats']}",
        "=" * 70,
        LOG_MESSAGES['total_execution'].format(time=total_time),
        LOG_MESSAGES['processed_actions'].format(count=processed_actions_count or 0),
        LOG_MESSAGES['executed_functions'].format(count=len(function_execution_times or {})),
        "",
        LOG_MESSAGES['execution_times'],
    ]
    
    # Добавление времени выполнения каждой функции
    if function_execution_times:
        for func_name, exec_time in function_execution_times.items():
            summary_lines.append(f"  - {func_name}: {exec_time:.4f} сек")
    else:
        summary_lines.append("  - Нет данных о времени выполнения функций")
    
    # Завершающие строки
    summary_lines.extend([
        "",
        LOG_MESSAGES['program_completed'].format(time=current_time),
        "=" * 70
    ])
    
    # Объединение в одну строку
    summary_text = "\n".join(summary_lines)
    
    # Вывод в консоль и лог
    if logger:
        logger.info(LOG_MESSAGES['summary_output'].format(summary=summary_text))
        logger.info(LOG_MESSAGES['summary_title'])
        logger.info(LOG_MESSAGES['total_time'].format(time=total_time) + f", {LOG_MESSAGES['actions_processed'].format(count=processed_actions_count or 0)}, {LOG_MESSAGES['functions_executed'].format(count=len(function_execution_times or {}))}")
        
        # Логирование времени каждой функции
        if function_execution_times:
            for func_name, exec_time in function_execution_times.items():
                logger.info(LOG_MESSAGES['function_time'].format(func=func_name, time=exec_time))

# =============================================================================
# ОСНОВНАЯ ПРОГРАММА
# =============================================================================

def main():
    """
    Основная функция программы
    
    Координирует выполнение всех этапов:
    1. Инициализация логирования
    2. Получение данных
    3. Генерация скриптов согласно настройкам
    4. Вывод статистики
    """
    global program_start_time
    
    # Инициализация времени начала программы
    program_start_time = time.time()
    
    # Настройка логирования
    setup_logging()
    
    # Инициализация основного логгера
    main_logger = get_script_logger("main", "execution")
    
    # Стартовое сообщение с разделителями для читаемости
    start_time_str = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]
    main_logger.info(LOG_MESSAGES['separator_line'])
    main_logger.info(LOG_MESSAGES['program_start'].format(time=start_time_str))
    main_logger.info(LOG_MESSAGES['processing_start_time'].format(time=start_time_str))
    main_logger.info(LOG_MESSAGES['logging_level'].format(level=LOG_LEVEL))
    main_logger.info(LOG_MESSAGES['separator_line'])
    
    try:
        # Выполнение операций для каждого активного скрипта
        if ACTIVE_SCRIPTS and len(ACTIVE_SCRIPTS) > 0:
            main_logger.info(f"Активные скрипты: {', '.join(ACTIVE_SCRIPTS)}")
            
            # ПЕРВЫЙ ЭТАП: Генерация всех скриптов
            main_logger.info(LOG_MESSAGES['stage1_title'])
            for script_name in ACTIVE_SCRIPTS:
                if script_name in FUNCTION_CONFIGS:
                    config = FUNCTION_CONFIGS[script_name]
                    active_operations = config.get("active_operations", "scripts_only")
                    
                    main_logger.info(LOG_MESSAGES['script_processing'].format(script_name=script_name))
                    main_logger.info(LOG_MESSAGES['active_operations_info'].format(script_name=script_name, operations=active_operations))
                    
                    # Генерация скриптов
                    if active_operations in ["scripts_only", "both"]:
                        main_logger.info(LOG_MESSAGES['script_generation_info'].format(script_name=script_name))
                        if script_name == "leaders_for_admin":
                            generate_leaders_for_admin_script()
                        elif script_name == "reward":
                            generate_reward_script()
                        elif script_name == "reward_profiles":
                            # reward_profiles теперь обрабатывается как часть reward
                            main_logger.info(LOG_MESSAGES['script_generation_skipped'].format(script_name=script_name, operations="внутри reward"))
                        elif script_name == "profile":
                            generate_profile_script()
                        elif script_name == "news_details":
                            generate_news_details_script()
                        elif script_name == "address_book_tn":
                            generate_address_book_tn_script()
                        elif script_name == "address_book_dev":
                            generate_address_book_dev_script()
                        elif script_name == "orders":
                            generate_orders_script()
                        elif script_name == "news_list":
                            generate_news_list_script()
                        elif script_name == "rating_list":
                            generate_rating_list_script()
                        else:
                            main_logger.warning(f"Неизвестный скрипт: {script_name}")
                    else:
                        main_logger.info(LOG_MESSAGES['script_generation_skipped'].format(script_name=script_name, operations=active_operations))
                else:
                    main_logger.warning(f"Скрипт '{script_name}' не найден в конфигурации FUNCTION_CONFIGS")
            
            # ВТОРОЙ ЭТАП: Обработка всех JSON файлов в Excel
            main_logger.info(LOG_MESSAGES['stage2_title'])
            for script_name in ACTIVE_SCRIPTS:
                if script_name in FUNCTION_CONFIGS:
                    config = FUNCTION_CONFIGS[script_name]
                    active_operations = config.get("active_operations", "scripts_only")
                    
                    # Обработка JSON файлов
                    if active_operations in ["json_only", "both"]:
                        # Проверяем вложенные конфигурации
                        if script_name == "reward" and "reward_profiles" in config:
                            # Обработка reward_profiles как части reward
                            reward_profiles_config = config["reward_profiles"]
                            if "json_file" in reward_profiles_config:
                                json_file = reward_profiles_config["json_file"]
                                main_logger.info(LOG_MESSAGES['json_file_processing_info'].format(json_file=json_file))
                                
                                convert_specific_json_file(json_file, "reward_profiles")
                            else:
                                main_logger.warning(f"Для скрипта {script_name} не указан json_file в конфигурации reward_profiles")
                        elif script_name == "leaders_for_admin" and "leaders_processing" in config:
                            # Обработка leaders_processing как части leaders_for_admin
                            leaders_processing_config = config["leaders_processing"]
                            if "json_file" in leaders_processing_config:
                                json_file = leaders_processing_config["json_file"]
                                main_logger.info(LOG_MESSAGES['json_file_processing_info'].format(json_file=json_file))
                                
                                convert_specific_json_file(json_file, script_name)
                            else:
                                main_logger.warning(f"Для скрипта {script_name} не указан json_file в конфигурации leaders_processing")
                        elif "json_file" in config:
                            # Прямая конфигурация json_file
                            json_file = config["json_file"]
                            main_logger.info(LOG_MESSAGES['json_file_processing_info'].format(json_file=json_file))
                            
                            convert_specific_json_file(json_file, script_name)
                        else:
                            main_logger.warning(f"Для скрипта {script_name} не указан json_file в конфигурации")
                    else:
                        main_logger.info(LOG_MESSAGES['json_processing_skipped'].format(script_name=script_name, operations=active_operations))
        else:
            main_logger.warning(LOG_MESSAGES['no_active_scripts'])
            
        # Альтернативный способ - ручной вызов конкретных функций
        # Раскомментируйте нужные строки для тестирования
        # generate_leaders_for_admin_script()  # CSV с разделителем ;
        # generate_profile_script()  # TXT с различными разделителями
        # generate_news_list_script()  # использует переменную согласно конфигурации
        
        main_logger.info(LOG_MESSAGES['program_success'])
        
    except Exception as e:
        # Обработка критических ошибок
        main_logger.error(LOG_MESSAGES['critical_error'].format(error=str(e)))
        
    finally:
        # Вывод итоговой статистики (всегда выполняется)
        print_summary()
        
        # Финальное сообщение
        end_time_str = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]
        total_time = time.time() - program_start_time
        main_logger.info(LOG_MESSAGES['separator_line'])
        main_logger.info(LOG_MESSAGES['program_end'].format(time=end_time_str))
        main_logger.info(LOG_MESSAGES['total_execution_time'].format(time=total_time))
        main_logger.info(LOG_MESSAGES['separator_line'])

# Точка входа в программу
if __name__ == "__main__":
    main() 